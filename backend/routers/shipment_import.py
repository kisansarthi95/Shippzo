"""
Shipment Import System — Phase F6.0 (2026-06).

Bulk update / create Shipments via user-uploaded CSV / XLSX files.

Unlike routers/file_import.py (which lands rows into `pending_orders`),
THIS router matches by Tracking Number against EXISTING shipments and
UPDATES them in place. Three import modes are supported:

    booking       — bulk update booking metadata (courier, order id,
                    items, customer, address). Cross-verify weight,
                    payment_mode, and COD amount → mismatch report.
    delivery      — mark parcels Delivered with delivered_at + POD.
    cod_payment   — record COD amount actually collected, remittance
                    date, and payer name.

Behaviour rules (per user requirement):
  • Match strictly by tracking_id (case-insensitive, trimmed).
  • If NOT matched → row is added to the "unmatched" downloadable report.
    We do NOT auto-create shipments here (that's what pending_orders
    imports are for).
  • Cross-verify fields (weight, payment_mode, cod_amount) between the
    IMPORTED row and the EXISTING shipment. If they DIFFER → we KEEP
    the existing value untouched AND log both values in the mismatch
    report. User can download the report and reconcile manually.
  • Fields that are NOT cross-verify targets are updated in place.

Endpoints:
    POST /api/shipments/import/preview             parse + suggest mapping
    POST /api/shipments/import/commit              validate + apply
    GET  /api/me/shipment-import-mapping           saved mapping (per type)
    PUT  /api/me/shipment-import-mapping           persist mapping
    GET  /api/shipments/import/batches             list past import batches
    GET  /api/shipments/import/batches/{id}        detail (per-row status)
    GET  /api/shipments/import/batches/{id}/mismatches.csv
                                                   downloadable mismatch CSV

Pattern: late-binding `init()` — same as routers/file_import.py.
"""
from __future__ import annotations

import csv
import io
import json
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, Query
from fastapi.responses import Response
from pydantic import BaseModel

from import_schema import (
    SCHEMA_FIELDS,
    NUMERIC_FIELDS,
    HEADER_ALIASES,
    suggest_mapping,
    normalise_value,
    normalise_status,
    is_custom_field_mapping,
    custom_field_id,
    validate_mapping_field,
)


shipment_import_router = APIRouter(prefix="/api", tags=["shipment-import"])


IMPORT_TYPES = {"booking", "delivery", "cod_payment"}

# Per-type target field allowlist. Any mapping outside these is
# ignored during commit so a user can't accidentally overwrite the
# customer name via a Delivery upload, for example.
TARGET_FIELDS_BY_TYPE: Dict[str, List[str]] = {
    "booking": [
        # Identity keys the merchant may want to sync (e.g. corrected
        # names / phones from the courier's manifest).
        "tracking_id",          # required — match key
        "order_id",
        "customer_name",
        "customer_phone",
        "customer_alt_phone",
        "customer_email",
        "address",
        "city",
        "state",
        "pincode",
        "items",
        "courier_hint",
        "notes",
        # Cross-verify fields (never override existing value on mismatch)
        "weight",
        "payment_mode",
        "amount",
    ],
    "delivery": [
        "tracking_id",           # required — match key
        "status",                # optional; auto-forced to "Delivered"
        "delivered_at",
        "pod_reference",
        "notes",
        # Phase F6.4 — free-text "Last Event" column from courier remit
        # sheet. Stored raw + classified into a short category for the
        # card badge (see classify_last_event()).
        "last_event",
        # Phase F7.6 (Jun-2026) — courier remit sheets frequently
        # carry a "Booked On" / "Booking Date" column. This is the
        # SAME `booking_date` DB column set by the Booking import.
        # Delivery import may only FILL the field when it is empty —
        # never overwrite an existing booking date (see commit path).
        "booking_date",
        # Still cross-verified even for delivery import
        "weight",
        "payment_mode",
        "amount",
    ],
    "cod_payment": [
        "tracking_id",           # required — match key
        "cod_collected_amount",
        "cod_payment_date",
        "cod_payer_name",
        "notes",
        # Cross-verify: the courier's reported COD amount must match
        # what we booked. If they differ, alert the merchant so they
        # can chase the courier before remittance.
        "amount",
        "payment_mode",
    ],
}

# Human labels for the UI dropdown.
FIELD_LABELS: Dict[str, str] = {
    "tracking_id":          "Tracking Number",
    "order_id":             "Order ID",
    "customer_name":        "Customer Name",
    "customer_phone":       "Customer Phone",
    "customer_alt_phone":   "Alternate Phone",
    "customer_email":       "Customer Email",
    "address":              "Address",
    "city":                 "City",
    "state":                "State",
    "pincode":              "Pincode",
    "items":                "Items",
    "courier_hint":         "Courier (Name)",
    "notes":                "Notes / Remarks",
    "weight":               "Weight",
    "payment_mode":         "Payment Mode (COD/PAID)",
    "amount":               "Amount / COD Amount",
    "status":               "Status",
    "delivered_at":         "Delivered On",
    "pod_reference":        "POD Reference / Receiver",
    "cod_collected_amount": "COD Amount Collected",
    "cod_payment_date":     "COD Payment / Remit Date",
    "cod_payer_name":       "COD Payer Name",
    "last_event":           "Last Event (India Post/Courier)",
    # Phase F7.6 — mappable in Delivery Update; uses the SAME
    # `booking_date` DB column set at booking-import time. Delivery
    # import will only fill it when currently empty; existing values
    # are preserved (never overwritten by remit-sheet data).
    "booking_date":         "Booking Date",
}


# Fields where a mismatch between imported and stored value must be
# recorded WITHOUT overriding the stored value. Applies across all
# import types (user requirement — "cross-verify, don't override").
CROSS_VERIFY_FIELDS: List[str] = ["weight", "payment_mode", "amount"]


# ────────────────────── constants ─────────────────────
MAX_SAMPLE_ROWS = 10
MAX_FILE_BYTES  = 10 * 1024 * 1024          # 10 MB hard limit (per user)
MAX_IMPORT_ROWS = 10000                     # generous — 10 MB cap prevails


# ────────────────────── payloads ──────────────────────
class MappingPayload(BaseModel):
    mapping: Dict[str, str]                 # { csv_col: schema_field }


# ════════════════════════════════════════════════════════════════════
#                              Parsers
# ════════════════════════════════════════════════════════════════════
#
# All parsers return a **raw grid** (List[List[Any]]) with NO header
# assumption. `_apply_layout()` then slices out the header row + data
# rows based on the user-provided `header_row`, `data_start_row`, and
# `header_col` values. This matches the Webhook Mapping Engine UX where
# spreadsheets with title banners / metadata before the actual header
# (common in courier remit sheets) can be imported cleanly.
# ════════════════════════════════════════════════════════════════════

MAX_RAW_ROWS = 12000    # safety ceiling — 10 MB CSV rarely exceeds this


def _parse_csv_raw(blob: bytes) -> List[List[Any]]:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = blob.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(status_code=400, detail="Could not decode file")
    reader = csv.reader(io.StringIO(text))
    grid: List[List[Any]] = []
    for line in reader:
        grid.append([(c or "").strip() if isinstance(c, str) else c for c in line])
        if len(grid) > MAX_RAW_ROWS:
            break
    if not grid:
        raise HTTPException(status_code=400, detail="File is empty")
    return grid


def _parse_xlsx_raw(blob: bytes) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="XLSX parser (openpyxl) not installed on the server",
        )
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read XLSX: {e}")
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Workbook has no sheets")
    grid: List[List[Any]] = []
    for line in ws.iter_rows(values_only=True):
        grid.append(list(line))
        if len(grid) > MAX_RAW_ROWS:
            break
    if not grid:
        raise HTTPException(status_code=400, detail="Sheet is empty")
    return grid


def _apply_layout(
    grid: List[List[Any]],
    header_row: int,
    data_start_row: int,
    header_col: int,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Slice a raw grid into (columns, list-of-row-dicts) using the
    1-based header_row / data_start_row / header_col config.

    • header_row     — 1-based row number where COLUMN HEADERS live.
    • data_start_row — 1-based row number where actual data begins.
                       Must be >= header_row + 1.
    • header_col     — 1-based column number where the FIRST meaningful
                       column lives. Any columns to the LEFT of this
                       are ignored entirely (both in header + data).

    Blank rows in the data region are skipped. Data columns beyond the
    header width are truncated; short rows are padded with "".
    """
    if header_row < 1:
        header_row = 1
    if header_col < 1:
        header_col = 1
    if data_start_row <= header_row:
        data_start_row = header_row + 1

    if header_row > len(grid):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Header row {header_row} is beyond the file "
                f"(file has {len(grid)} rows)."
            ),
        )
    header_line = grid[header_row - 1]
    header_line = header_line[header_col - 1:]
    columns: List[str] = []
    for c in header_line:
        if c is None:
            columns.append("")
        elif isinstance(c, str):
            columns.append(c.strip())
        else:
            columns.append(str(c).strip())
    # Drop trailing blank header columns — they add noise to the mapping UI.
    while columns and not columns[-1]:
        columns.pop()
    if not columns:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Header row {header_row} is empty (no column names found). "
                "Check the Header Row / Data Start Column values."
            ),
        )

    rows: List[Dict[str, Any]] = []
    for row_i in range(data_start_row - 1, len(grid)):
        line = grid[row_i]
        line = line[header_col - 1:]
        # Pad or trim to header width.
        if len(line) < len(columns):
            line = list(line) + [None] * (len(columns) - len(line))
        else:
            line = line[: len(columns)]
        # Skip fully-blank rows.
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in line):
            continue
        rows.append({
            columns[i]: (
                line[i].strip() if isinstance(line[i], str) else line[i]
            )
            for i in range(len(columns))
        })
    return columns, rows


def _parse_csv(blob: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    """Legacy default-layout parser: header on row 1, data from row 2.
    Retained so /orders/import/* + tests keep working unchanged."""
    grid = _parse_csv_raw(blob)
    cols, rows = _apply_layout(grid, header_row=1, data_start_row=2, header_col=1)
    return cols, rows  # type: ignore[return-value]


def _parse_xlsx(blob: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Legacy default-layout parser."""
    grid = _parse_xlsx_raw(blob)
    return _apply_layout(grid, header_row=1, data_start_row=2, header_col=1)


def _parse_upload(
    file: UploadFile,
    blob: bytes,
    header_row: int = 1,
    data_start_row: int = 2,
    header_col: int = 1,
) -> Tuple[str, List[str], List[Dict[str, Any]]]:
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        grid = _parse_xlsx_raw(blob)
        cols, rows = _apply_layout(grid, header_row, data_start_row, header_col)
        return "xlsx", cols, rows
    if name.endswith(".csv") or (file.content_type or "").startswith("text/csv"):
        grid = _parse_csv_raw(blob)
        cols, rows = _apply_layout(grid, header_row, data_start_row, header_col)
        return "csv", cols, rows
    raise HTTPException(
        status_code=400,
        detail=("Only .csv and .xlsx files are supported (got "
                f"{file.filename or 'unknown'})"),
    )


def _row_sample(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    for r in rows[:MAX_SAMPLE_ROWS]:
        out.append({
            k: ("" if v is None else (
                v.isoformat() if hasattr(v, "isoformat") else str(v).strip()
            ))
            for k, v in r.items()
        })
    return out


def _raw_preview_grid(
    grid: List[List[Any]],
    max_rows: int = 12,
    max_cols: int = 20,
) -> List[List[str]]:
    """Return the first N rows × M cols of the raw grid, stringified
    for the UI 'first 12 rows peek' block. Lets the user visually
    confirm which row IS the header before finalising the layout."""
    out: List[List[str]] = []
    for r in grid[:max_rows]:
        r2 = r[:max_cols]
        out.append([
            "" if c is None else (
                c.isoformat() if hasattr(c, "isoformat") else str(c).strip()
            )
            for c in r2
        ])
    return out


# ════════════════════════════════════════════════════════════════════
#                       Weight normalisation
# ════════════════════════════════════════════════════════════════════

_WEIGHT_RE = re.compile(r"([-+]?\d*\.?\d+)\s*(kg|kgs|g|gm|gms|gram|grams|kilogram|kilograms)?", re.I)


def _weight_to_grams(raw: Any) -> Optional[float]:
    """Best-effort conversion of a free-text weight into grams.

    Returns None when input is empty or unparseable. Used ONLY for
    the cross-verify comparison (a fuzzy 1g tolerance is applied by
    the caller). The stored `weight` field always keeps the original
    formatted string; this helper doesn't mutate anything.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = _WEIGHT_RE.search(s)
    if not m:
        return None
    try:
        val = float(m.group(1))
    except (TypeError, ValueError):
        return None
    unit = (m.group(2) or "").lower()
    if unit.startswith("kg") or unit.startswith("kilogram"):
        return val * 1000.0
    # Default to grams when no unit or 'g' family.
    return val


def _payment_mode_equal(a: Any, b: Any) -> bool:
    """Case-insensitive comparison after normalise_value."""
    an = normalise_value("payment_mode", a) if a not in (None, "") else ""
    bn = normalise_value("payment_mode", b) if b not in (None, "") else ""
    return (an or "").strip().upper() == (bn or "").strip().upper()


def _amount_equal(a: Any, b: Any, tol: float = 0.5) -> bool:
    try:
        af = float(a or 0.0)
        bf = float(b or 0.0)
    except (TypeError, ValueError):
        return False
    return abs(af - bf) <= tol


# ─── Phase F11.A — Junk tracking-value detector ─────────────────────
#
# Courier remit spreadsheets very frequently carry summary rows like
# "Total", "Sub Total", "Grand Total", "RS 12345", "N/A" mixed into
# the tracking-id column. Without a filter, every one of those rows
# lands in the batch tally as an "unmatched" article, polluting the
# Article Number Mismatch report the operator uses to chase courier
# discrepancies.
#
# `_is_junk_tracking_value()` returns True for cells that look like
# spreadsheet noise and should be skipped SILENTLY (they never touch
# the unmatched/error counters). Deliberately conservative — any real
# tracking-id-looking value (alphanumeric ≥5 chars) passes through.
_JUNK_TOKENS = {
    "total", "totals",
    "sub total", "subtotal", "sub-total",
    "grand total", "grand-total",
    "amount total", "total amount",
    "rs", "rs.", "rupees", "inr", "₹",
    "sum", "sum total",
    "n/a", "na", "nil", "none", "null",
    "-", "--", "—", "–",
    "0", "0.0", "0.00",
    # Common column-header echoes that leak into data rows when a
    # sheet has multiple header bands.
    "tracking id", "tracking no", "tracking number",
    "awb", "awb no", "awb number",
    "article no", "article number", "consignment", "consignment no",
    "order id", "order no",
}
_JUNK_TOKEN_PATTERNS = (
    re.compile(r"^rs\.?\s*[0-9,\.]+$", re.I),           # "RS 1200" / "Rs.1,200"
    re.compile(r"^total\s*[:\-]?\s*[0-9,\.]*$", re.I),  # "Total 5", "Total:"
    re.compile(r"^grand\s*total\b.*$", re.I),           # "Grand Total 8500"
    re.compile(r"^sub\s*total\b.*$", re.I),
    re.compile(r"^amount\s*[:\-]?\s*[0-9,\.]+$", re.I), # "Amount: 500"
    re.compile(r"^page\s+\d+", re.I),                   # "Page 1 of 3"
)


def _is_junk_tracking_value(raw: Any) -> bool:
    """Detect spreadsheet-junk cells masquerading as a tracking id.

    Rules:
      • Blank / whitespace-only  → junk
      • Value in _JUNK_TOKENS    → junk
      • Matches _JUNK_TOKEN_PATTERNS → junk
      • Pure whitespace or symbols only → junk
      • Anything else            → real tracking value (let it flow)
    """
    if raw is None:
        return True
    s = str(raw).strip()
    if not s:
        return True
    low = s.lower()
    if low in _JUNK_TOKENS:
        return True
    for pat in _JUNK_TOKEN_PATTERNS:
        if pat.match(s):
            return True
    # Non-alphanumeric only? (e.g. "***", "---", "===")
    if not re.search(r"[A-Za-z0-9]", s):
        return True
    return False


# ─── Phase F6.4 — Last Event Category classifier ────────────────────
# India Post / courier remit sheets carry a free-text "Last Event"
# column like "Item Kept on Hold at Wanri B.O on 09/07/2026 13:04:10".
# The Shipments card only has room for a short badge, so we map the
# raw text to ONE of these fixed categories. Shipment Details keeps
# the ORIGINAL text untouched. Order matters — more specific patterns
# come first so "Out for Delivery" doesn't accidentally hit "Delivered".
#
# Phase F7.7 (Jun-2026) — added specific India Post phrasings so the
# UI can show the actual event name (not just the generic bucket) on
# the shipment card. New categories added:
#   • "Return Review" ← "Item Delivered(Sender)" / "Delivered to Sender"
#                       (RTS deliveries. Never auto-moves to Returned —
#                        needs user confirmation, see UI flow.)
#   • "Item Returned to Sender" ← "Item Returned to Sender" (priority)
#   • "Item Kept on Hold"       ← keeps the exact India Post label
#   • "Item Dispatched"         ← "Item Dispatched" specifically
#   • "Item Bagged"             ← "Item bagged"
#   • "Item Received"           ← "Item Received"
_LAST_EVENT_PATTERNS: List[Tuple[str, str]] = [
    # Return-to-Sender DELIVERY event — MUST match before plain
    # "delivered" or the RTS acknowledgement gets absorbed into
    # the Delivered bucket.
    (r"deliver(?:ed|y)?\s*[\(\-\/\s]*sender",         "Return Review"),
    (r"\bdelivered\b",                                "Delivered"),
    (r"\bout\s*for\s*delivery\b|\bofd\b",             "Out for Delivery"),
    # Priority OFD variants — kept as their exact India Post label so
    # the card badge reads verbatim.
    (r"return(?:ed)?\s*to\s*sender|\brts\b",          "Item Returned to Sender"),
    (r"\bkept\s*on\s*hold\b|\bon\s*hold\b|\bheld\b",  "Item Kept on Hold"),
    (r"\bhold\b|\bmisc\b",                            "Item Kept on Hold"),
    (r"\bredirect(?:ed|ion)?\b|\brerout",             "Redirected"),
    (r"\breturn(?:ed|ing)?\b",                        "Returned"),
    # In-Transit signals (India Post uses these three verbatim).
    (r"\bitem\s*dispatched\b|\bdispatch(?:ed|ing)?\b|\bin\s*transit\b",
                                                      "Item Dispatched"),
    (r"\bitem\s*bag(?:ged)?\b|\bbag(?:ged)?\b|\bin\s*bag\b",
                                                      "Item Bagged"),
    (r"\bitem\s*receiv(?:ed|ing)?\b|\breceiv(?:ed|ing)?\b|\bbooked\b",
                                                      "Item Received"),
]
_LAST_EVENT_COMPILED = [(re.compile(p, re.I), cat) for p, cat in _LAST_EVENT_PATTERNS]


def classify_last_event(text: Any) -> str:
    """Return the short category badge label for a raw Last Event
    string. Never modifies the input; returns "" for empty input and
    "Other" when no known pattern matches."""
    if not text:
        return ""
    s = str(text).strip()
    if not s:
        return ""
    for rx, cat in _LAST_EVENT_COMPILED:
        if rx.search(s):
            return cat
    return "Other"


# ════════════════════════════════════════════════════════════════════
# Phase F8.0 — SHARED Shipment Update Engine.
#
# One single place that applies a "delivery event" (raw Last-Event
# text) to a shipment. Used by BOTH:
#   • the Delivery Import commit loop (source="import"), and
#   • the Courier-Sync SMS ingest pipeline (source="sms",
#     routers/courier_sync.py).
# so the two paths can never drift apart on business rules.
#
# Status routing (F7.7):
#   • Delivered                → Delivered  (+ delivered_at,
#                                confirmation_status="confirmed")
#   • Return Review            → NO status change; sets
#                                needs_return_review flag
#   • Item Dispatched/Bagged/
#     Received                 → In Transit
#   • everything else known    → Out for Delivery
#
# SMS specifics (source="sms"):
#   • status comes from the courier's configured Scanning Rules
#     (status_override) — NOT from the route table — so the operator's
#     per-courier keyword→stage config remains the single authority.
#   • `category_fallback` = canonical status from the matched rule.
#     Used when the raw SMS doesn't classify (e.g. "is Successful"
#     carries no known event words) so badges & confirmations still
#     fire per the configured rule.
#   • booking_date is stamped from the SMS ONLY when currently empty.
#   • delivered_at prefers the datetime extracted from the SMS.
# ════════════════════════════════════════════════════════════════════

STATUS_ROUTE_BY_EVENT: Dict[str, str] = {
    "Delivered":                "Delivered",
    "Out for Delivery":         "Out for Delivery",
    "Item Returned to Sender":  "Out for Delivery",
    "Item Kept on Hold":        "Out for Delivery",
    "Item Dispatched":          "In Transit",
    "Item Bagged":              "In Transit",
    "Item Received":            "In Transit",
    "Redirected":               "Out for Delivery",
    "Returned":                 "Out for Delivery",
    "Other":                    "Out for Delivery",
    # "Return Review" intentionally absent — never auto-routes.
}


def apply_last_event_engine(
    existing: Dict[str, Any],
    applied: Dict[str, Any],
    *,
    raw_event: str,
    now: str,
    source: str = "import",              # "import" | "sms"
    status_override: Optional[str] = None,  # sms: rule-driven status
    category_fallback: str = "",            # sms: canonical from rule
    event_dt: str = "",                     # sms: datetime parsed from SMS
) -> str:
    """Apply one delivery/courier event to the `applied` update dict.

    Mutates `applied` in place (mirrors how the import commit loop
    builds its per-row update) and returns the final event category.
    `existing` is the current shipment document (read-only).
    """
    raw_ev = (raw_event or "").strip()
    ev_cat = classify_last_event(raw_ev) if raw_ev else ""

    if source == "sms":
        # SMS body IS the event — persist it so Shipment Details and
        # the card badge behave exactly like a Delivery Import row.
        if raw_ev:
            applied["last_event"] = raw_ev
        # Fall back to the Scanning-Rule canonical when the raw text
        # doesn't classify (keeps rule config as the authority).
        if (not ev_cat or ev_cat == "Other") and category_fallback:
            ev_cat = category_fallback

    if ev_cat:
        applied["last_event_category"] = ev_cat

    # ── Status routing ───────────────────────────────────────────
    if source == "import":
        if ev_cat and ev_cat != "Return Review":
            applied["status"] = STATUS_ROUTE_BY_EVENT.get(ev_cat, "Out for Delivery")
        elif not ev_cat and applied.get("delivered_at"):
            # F6.0 contract — a delivery row carrying an explicit
            # delivery date but NO event text IS a delivered row.
            applied["status"] = "Delivered"
            applied["confirmation_status"] = "confirmed"
    else:
        # SMS: the courier's Scanning Rule decides the status.
        # Return Review NEVER auto-moves (F7.7 policy) even if the
        # rule keyword also matched.
        if status_override and ev_cat != "Return Review":
            applied["status"] = status_override

    # ── Return Review flag (both paths) ──────────────────────────
    if ev_cat == "Return Review":
        applied["needs_return_review"] = True
        applied["return_review_at"] = now

    # ── delivered_at stamp ────────────────────────────────────────
    if source == "import":
        if applied.get("status") == "Delivered" and not applied.get("delivered_at"):
            applied["delivered_at"] = now
    else:
        delivered_now = ev_cat != "Return Review" and (
            applied.get("status") == "Delivered"
            or ev_cat == "Delivered"
            or category_fallback == "Delivered"
        )
        if (
            delivered_now
            and not existing.get("delivered_at")
            and not applied.get("delivered_at")
        ):
            applied["delivered_at"] = event_dt or now

    # ── Delivery Confirmation — auto-confirm ONLY on Delivered. ──
    if ev_cat == "Delivered":
        applied["confirmation_status"] = "confirmed"

    # ── Booking SMS → booking_date, ONLY if currently empty. ─────
    if (
        source == "sms"
        and category_fallback == "Booked"
        and event_dt
        and not existing.get("booking_date")
    ):
        applied["booking_date"] = event_dt

    return ev_cat


def _weight_equal(a: Any, b: Any) -> bool:
    ga = _weight_to_grams(a)
    gb = _weight_to_grams(b)
    if ga is None and gb is None:
        return True
    if ga is None or gb is None:
        return False
    # 1g tolerance — courier scales round differently than merchant scales.
    return abs(ga - gb) <= 1.0


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ════════════════════════════════════════════════════════════════════
#                              Router
# ════════════════════════════════════════════════════════════════════

def init() -> None:
    """Register routes after server.py finishes initialising."""
    import logging
    _logger = logging.getLogger("routers.shipment_import")
    from server import (  # noqa: WPS433 — late binding
        db,
        get_current_user as _get_current_user,
    )

    # ── Manual Correction (Phase F10.3+) ──────────────────────────
    #
    # Lets the operator fix a COD amount mismatch without re-uploading
    # the entire courier remit file. The user opens Shipment Details
    # for a flagged shipment, taps "Manual Correction", enters the
    # correct received amount and (optionally) the payer name, and
    # this endpoint:
    #   • updates cod_collected_amount on the shipment
    #   • sets cod_payment_status="received" + cod_payment_date if
    #     the row was still pending
    #   • drops any payment/cod_amount entries from
    #     import_validation_alerts if the new received amount matches
    #     the booked cod_amount (within tolerance). Non-payment
    #     alerts (weight, payment_mode) are preserved.
    class _ManualCodCorrection(BaseModel):
        cod_collected_amount: float
        cod_payer_name: Optional[str] = None
        cod_payment_date: Optional[str] = None  # ISO — optional override

    @shipment_import_router.post("/shipments/{shipment_id}/manual-cod-correction")
    async def manual_cod_correction(
        shipment_id: str,
        payload: _ManualCodCorrection,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        ship = await db.shipments.find_one(
            {"id": shipment_id, "user_id": current_user["id"]},
            {"_id": 0},
        )
        if not ship:
            raise HTTPException(status_code=404, detail="Shipment not found")

        received = float(payload.cod_collected_amount or 0.0)
        if received < 0:
            raise HTTPException(status_code=400, detail="Amount cannot be negative")

        booked = ship.get("cod_amount")
        now = _iso_now()

        set_fields: Dict[str, Any] = {
            "cod_collected_amount": received,
            "cod_payment_status": "received",
            "last_import_at": now,
            "last_import_type": "manual_correction",
            "modified_at": now,
        }
        if not ship.get("cod_payment_date"):
            set_fields["cod_payment_date"] = payload.cod_payment_date or now
        elif payload.cod_payment_date:
            set_fields["cod_payment_date"] = payload.cod_payment_date
        if payload.cod_payer_name is not None:
            set_fields["cod_payer_name"] = payload.cod_payer_name

        # Prune payment-related alerts from import_validation_alerts
        # when the correction resolves the discrepancy. Preserve
        # non-payment alerts (weight / payment_mode) untouched.
        alerts = list(ship.get("import_validation_alerts") or [])
        payment_fields = {"cod_amount", "amount"}
        if alerts:
            if booked is not None and _amount_equal(booked, received):
                kept = [a for a in alerts if a.get("field") not in payment_fields]
                set_fields["import_validation_alerts"] = kept
            else:
                # Refresh the payment alert entries to reflect the new
                # received value so the UI shows the current delta.
                new_alerts = []
                had_payment_alert = False
                for a in alerts:
                    if a.get("field") in payment_fields:
                        had_payment_alert = True
                        new_alerts.append({
                            "field":    "cod_amount",
                            "existing": booked,
                            "imported": received,
                            "at":       now,
                        })
                    else:
                        new_alerts.append(a)
                if not had_payment_alert and booked is not None:
                    new_alerts.append({
                        "field":    "cod_amount",
                        "existing": booked,
                        "imported": received,
                        "at":       now,
                    })
                set_fields["import_validation_alerts"] = new_alerts
        elif booked is not None and not _amount_equal(booked, received):
            # No prior alerts but the manual correction creates a
            # mismatch — seed one so the UI stays accurate.
            set_fields["import_validation_alerts"] = [{
                "field":    "cod_amount",
                "existing": booked,
                "imported": received,
                "at":       now,
            }]

        await db.shipments.update_one(
            {"id": shipment_id, "user_id": current_user["id"]},
            {"$set": set_fields},
        )
        fresh = await db.shipments.find_one(
            {"id": shipment_id, "user_id": current_user["id"]}, {"_id": 0},
        )
        return {
            "ok": True,
            "shipment": fresh,
            "matched": booked is not None and _amount_equal(booked, received),
        }

    # ── Saved default mapping (per user × import_type) ────────────

    @shipment_import_router.get("/me/shipment-import-mapping")
    async def get_saved_mapping(
        import_type: str = Query("booking", pattern="^(booking|delivery|cod_payment)$"),
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        doc = await db.users.find_one(
            {"id": current_user["id"]},
            {"_id": 0, "shipment_import_mappings": 1, "shipment_import_layouts": 1},
        ) or {}
        all_maps: Dict[str, Any] = doc.get("shipment_import_mappings") or {}
        all_layouts: Dict[str, Any] = doc.get("shipment_import_layouts") or {}
        layout = all_layouts.get(import_type) or {}
        return {
            "import_type":    import_type,
            "mapping":        all_maps.get(import_type) or {},
            "header_row":     int(layout.get("header_row") or 1),
            "data_start_row": int(layout.get("data_start_row") or 2),
            "header_col":     int(layout.get("header_col") or 1),
            "target_fields":  [
                {"key": f, "label": FIELD_LABELS.get(f, f), "required": f == "tracking_id"}
                for f in TARGET_FIELDS_BY_TYPE[import_type]
            ],
            "cross_verify":   CROSS_VERIFY_FIELDS,
        }

    @shipment_import_router.put("/me/shipment-import-mapping")
    async def put_saved_mapping(
        payload: MappingPayload,
        import_type: str = Query("booking", pattern="^(booking|delivery|cod_payment)$"),
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        allowed = set(TARGET_FIELDS_BY_TYPE[import_type])
        bad = [
            v for v in payload.mapping.values()
            if v and v not in allowed and not is_custom_field_mapping(v)
        ]
        if bad:
            raise HTTPException(
                status_code=400,
                detail=f"Fields not allowed for {import_type}: {sorted(set(bad))}",
            )
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": {f"shipment_import_mappings.{import_type}": payload.mapping}},
        )
        return {"ok": True, "import_type": import_type, "mapping": payload.mapping}

    # ── Preview (no DB write) ─────────────────────────────────────

    @shipment_import_router.post("/shipments/import/preview")
    async def import_preview(
        file: UploadFile = File(...),
        import_type: str = Form(..., description="booking | delivery | cod_payment"),
        header_row: int = Form(1, description="1-based row where COLUMN HEADERS live"),
        data_start_row: int = Form(2, description="1-based row where DATA begins"),
        header_col: int = Form(1, description="1-based column where data begins (skip N-1 left columns)"),
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        if import_type not in IMPORT_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"import_type must be one of {sorted(IMPORT_TYPES)}",
            )
        blob = await file.read()
        if len(blob) > MAX_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large (limit { MAX_FILE_BYTES // (1024 * 1024) } MB)",
            )

        # Parse the raw grid ONCE — used both for the layout-aware
        # (columns, rows) view AND for the "first 12 rows peek" that
        # helps the user visually locate the correct header row.
        name = (file.filename or "").lower()
        if name.endswith(".xlsx"):
            grid = _parse_xlsx_raw(blob)
            fmt = "xlsx"
        elif name.endswith(".csv") or (file.content_type or "").startswith("text/csv"):
            grid = _parse_csv_raw(blob)
            fmt = "csv"
        else:
            raise HTTPException(
                status_code=400,
                detail=("Only .csv and .xlsx files are supported (got "
                        f"{file.filename or 'unknown'})"),
            )
        columns, rows = _apply_layout(grid, header_row, data_start_row, header_col)

        # Filter suggest_mapping output to only fields relevant for this
        # import type. Keeps the UI focussed.
        allowed = set(TARGET_FIELDS_BY_TYPE[import_type])
        saved_doc = await db.users.find_one(
            {"id": current_user["id"]},
            {"_id": 0, "shipment_import_mappings": 1, "shipment_import_layouts": 1},
        ) or {}
        saved = saved_doc.get("shipment_import_mappings") or {}
        saved_map = (saved.get(import_type) or {})
        raw_suggested = suggest_mapping(columns, saved_map, [])
        suggested = {c: f for c, f in raw_suggested.items() if f in allowed}

        # Quick preview stats — how many rows carry a plausible tracking id?
        tracking_col = None
        for c, f in suggested.items():
            if f == "tracking_id":
                tracking_col = c
                break
        tracking_count = 0
        if tracking_col:
            for r in rows:
                v = r.get(tracking_col)
                if v is not None and str(v).strip():
                    tracking_count += 1

        return {
            "import_type":    import_type,
            "format":         fmt,
            "filename":       file.filename or "",
            "columns":        columns,
            "sample_rows":    _row_sample(rows),
            "total_rows":     len(rows),
            "rows_with_tracking": tracking_count,
            "target_fields":  [
                {"key": f, "label": FIELD_LABELS.get(f, f), "required": f == "tracking_id"}
                for f in TARGET_FIELDS_BY_TYPE[import_type]
            ],
            "cross_verify":   CROSS_VERIFY_FIELDS,
            "suggested":      suggested,
            # Layout the parser used — echoed back so the UI can display
            # exactly what got applied.
            "header_row":     header_row,
            "data_start_row": data_start_row,
            "header_col":     header_col,
            # First 12 rows × 20 cols of the ORIGINAL grid (pre-layout)
            # so the UI can render a scrollable peek widget the user
            # taps to say "row 3 IS my header".
            "raw_preview":    _raw_preview_grid(grid, max_rows=12, max_cols=20),
            "raw_total_rows": len(grid),
        }

    # ── Commit (validate + apply) ─────────────────────────────────

    @shipment_import_router.post("/shipments/import/commit")
    async def import_commit(
        file: UploadFile = File(...),
        import_type: str = Form(..., description="booking | delivery | cod_payment"),
        mapping: str = Form(..., description="JSON string {csv_col: schema_field}"),
        save_default: bool = Form(False),
        header_row: int = Form(1),
        data_start_row: int = Form(2),
        header_col: int = Form(1),
        # Phase F6.3 — For import_type=cod_payment, an OPTIONAL payment
        # batch payload can be sent to group these settlements under a
        # PaymentBatch record (e.g. one cheque = one batch). Ignored for
        # other import types.
        payment_batch: str = Form("", description="JSON string with PaymentBatch fields (name, payment_date, payment_mode, reference_number, ...)"),
        override_duplicate: bool = Form(False, description="If true, proceed even when reference_number already exists"),
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        if import_type not in IMPORT_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"import_type must be one of {sorted(IMPORT_TYPES)}",
            )
        try:
            mapping_obj: Dict[str, str] = json.loads(mapping)
            if not isinstance(mapping_obj, dict):
                raise ValueError("mapping must be a JSON object")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid mapping JSON: {e}")

        allowed = set(TARGET_FIELDS_BY_TYPE[import_type])
        # Reject unknown target fields early.
        bad = [v for v in mapping_obj.values() if v and v not in allowed]
        if bad:
            raise HTTPException(
                status_code=400,
                detail=f"Fields not allowed for {import_type}: {sorted(set(bad))}",
            )

        # tracking_id column MUST be mapped — no match key means we
        # can't safely locate any existing shipment.
        field_to_col: Dict[str, str] = {}
        for col, field in mapping_obj.items():
            if not field:
                continue
            # Last mapping wins per field (allows the UI to remap).
            field_to_col[field] = col
        if "tracking_id" not in field_to_col:
            raise HTTPException(
                status_code=400,
                detail="Please map a column to 'Tracking Number' before importing.",
            )

        blob = await file.read()
        if len(blob) > MAX_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large (limit { MAX_FILE_BYTES // (1024 * 1024) } MB)",
            )
        fmt, columns, rows = _parse_upload(file, blob, header_row, data_start_row, header_col)
        if len(rows) > MAX_IMPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(f"Too many rows ({len(rows)}). Max {MAX_IMPORT_ROWS} per upload."),
            )

        if save_default:
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {
                    f"shipment_import_mappings.{import_type}": mapping_obj,
                    f"shipment_import_layouts.{import_type}": {
                        "header_row":     header_row,
                        "data_start_row": data_start_row,
                        "header_col":     header_col,
                    },
                }},
            )

        # Phase F6.3 — parse & validate payment_batch metadata (only
        # relevant for cod_payment imports). We create the batch AFTER
        # the shipment updates so total_articles / total_amount reflect
        # the real result.
        payment_batch_payload: Optional[Dict[str, Any]] = None
        duplicate_batch: Optional[Dict[str, Any]] = None
        if import_type == "cod_payment" and payment_batch and payment_batch.strip():
            try:
                pb_raw = json.loads(payment_batch)
                if pb_raw and isinstance(pb_raw, dict):
                    from routers.payment_batches import (
                        PaymentBatchIn, validate_payment_batch, find_duplicate_ref,
                    )
                    pb_in = PaymentBatchIn(**pb_raw)
                    payment_batch_payload = validate_payment_batch(pb_in)
                    duplicate_batch = await find_duplicate_ref(
                        db, current_user["id"], payment_batch_payload["reference_number"],
                    )
                    if duplicate_batch and not override_duplicate:
                        # Soft warning — return 409 so UI can show a
                        # confirm dialog and re-submit with
                        # override_duplicate=true.
                        raise HTTPException(
                            status_code=409,
                            detail={
                                "error": "duplicate_reference",
                                "message": (
                                    f"Reference '{payment_batch_payload['reference_number']}' "
                                    f"already used in batch '{duplicate_batch.get('name')}' on "
                                    f"{duplicate_batch.get('payment_date')}. "
                                    "Import blocked to prevent duplicate payment."
                                ),
                                "existing_batch": duplicate_batch,
                            },
                        )
            except HTTPException:
                raise
            except Exception as _pb_e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid payment_batch payload: {_pb_e}",
                )

        # Batch document container — persisted at the end.
        batch_id = str(uuid.uuid4())
        now = _iso_now()
        batch_doc: Dict[str, Any] = {
            "id":                batch_id,
            "user_id":           current_user["id"],
            "import_type":       import_type,
            "filename":          file.filename or "",
            "format":            fmt,
            "created_at":        now,
            "mapping_used":      mapping_obj,
            "layout_used": {
                "header_row":     header_row,
                "data_start_row": data_start_row,
                "header_col":     header_col,
            },
            "total_rows":        len(rows),
            "matched_updated":   0,
            "matched_mismatch":  0,
            "matched_no_change": 0,
            "unmatched":         0,
            "errors":            0,
            # Phase F11.A — Rows whose tracking-id cell was noise
            # (Total / RS / N-A / blank / etc). We skip them silently
            # so they never inflate the "unmatched" tally.
            "junk_skipped":      0,
            # Phase F7.6 — Delivery-import-only stats for the newly
            # mappable `booking_date` field. These counters are 0 for
            # booking / cod_payment imports and never surfaced in
            # their summaries.
            "booking_date_updated":  0,
            "booking_date_skipped":  0,
            "booking_date_invalid":  0,
            "rows":              [],
        }

        # Build a tracking → shipment lookup in ONE query per batch.
        # We collect ALL tracking values first so we do exactly one
        # Mongo find({tracking_id: {$in: […]}}) regardless of file size.
        tracking_col = field_to_col["tracking_id"]

        def _tracking_of(row: Dict[str, Any]) -> str:
            v = row.get(tracking_col)
            if v is None:
                return ""
            return str(v).strip()

        tracking_values = list({
            _tracking_of(r) for r in rows
            if _tracking_of(r) and not _is_junk_tracking_value(_tracking_of(r))
        })
        # Case-insensitive lookup — build a normalised map.
        existing_lookup: Dict[str, Dict[str, Any]] = {}
        if tracking_values:
            # Match on tracking_id OR manual_tracking_id OR order_id / master_order_id
            # so users can also feed a courier remit sheet that uses the merchant's
            # Master Order ID as the key when the courier hasn't printed AWBs.
            or_clauses: List[Dict[str, Any]] = []
            for f in ("tracking_id", "manual_tracking_id", "order_id", "master_order_id"):
                or_clauses.append({f: {"$in": tracking_values}})
            cursor = db.shipments.find(
                {"user_id": current_user["id"], "$or": or_clauses},
                {"_id": 0},
            )
            async for sh in cursor:
                for f in ("tracking_id", "manual_tracking_id", "order_id", "master_order_id"):
                    v = str(sh.get(f) or "").strip()
                    if v:
                        existing_lookup[v.lower()] = sh
                        existing_lookup[v] = sh

        # ────────────── Per-row processing ──────────────
        updates_to_write: List[Tuple[str, Dict[str, Any]]] = []
        for idx, row in enumerate(rows):
            tracking = _tracking_of(row)
            row_status: str = "unmatched"
            shipment_id = ""
            mismatches: List[Dict[str, Any]] = []
            applied: Dict[str, Any] = {}
            row_error = ""
            # Phase F7.6 — Per-row booking_date outcome (only meaningful
            # for delivery imports). One of:
            #   • ""        — field not mapped OR cell blank OR non-delivery.
            #   • "updated" — cell parsed, no existing value → wrote it.
            #   • "skipped" — cell parsed, existing value → left alone.
            #   • "invalid" — cell present, parser rejected → Invalid Date.
            booking_date_row_status: str = ""

            try:
                if not tracking:
                    row_status = "unmatched"
                elif _is_junk_tracking_value(tracking):
                    # Phase F11.A — spreadsheet noise. Skip silently so
                    # it never lands in the unmatched tally / article
                    # mismatch report. Marked with its own status so
                    # the batch drill-down UI can still surface these
                    # rows if the operator wants to double-check.
                    row_status = "junk_skipped"
                else:
                    existing = existing_lookup.get(tracking) or existing_lookup.get(tracking.lower())
                    if not existing:
                        row_status = "unmatched"
                    else:
                        shipment_id = existing.get("id") or ""

                        # Build the update dict for this row.
                        # ─── Phase F6.2 — Also stamp IMPORT-SPECIFIC fields
                        # on the shipment so Shipment Details screen can
                        # surface the imported courier weight, payment
                        # type, and booked COD amount WITHOUT overwriting
                        # the merchant's original values.
                        row_imported_weight = ""
                        row_imported_payment_mode = ""
                        row_imported_amount: Optional[float] = None

                        for field, col in field_to_col.items():
                            if field == "tracking_id":
                                continue
                            raw = row.get(col)
                            new_val = normalise_value(field, raw)
                            # ── Phase F7.6 — Delivery-import booking_date
                            # needs the "empty vs invalid" distinction which
                            # the generic blank-guard below would collapse.
                            # We check it here BEFORE that guard.
                            #   • cell empty              → no counter (Blank)
                            #   • cell present + parsed   → Updated / Skipped
                            #   • cell present + unparsed → Invalid Date
                            if field == "booking_date" and import_type == "delivery":
                                raw_str = "" if raw is None else str(raw).strip()
                                if not raw_str:
                                    # Truly blank cell → don't touch anything.
                                    continue
                                if not new_val:
                                    booking_date_row_status = "invalid"
                                    continue
                                if existing.get("booking_date"):
                                    booking_date_row_status = "skipped"
                                    continue
                                booking_date_row_status = "updated"
                                applied["booking_date"] = new_val
                                continue
                            if new_val in ("", 0.0, None):
                                # Blank cell → don't touch the existing value.
                                continue

                            # ── Cross-verify guard ──
                            if field in CROSS_VERIFY_FIELDS:
                                # Phase F10.2 — for cod_payment imports the
                                # "amount" baseline is the shipment's COD
                                # amount (what the courier is supposed to
                                # collect: 669), NOT the order total which
                                # already netted the customer's advance
                                # token (719). Using the wrong baseline was
                                # producing false matched_ok rows that
                                # never surfaced the payment discrepancy
                                # in the batch tally or the Shipment
                                # Details alert.
                                if field == "amount":
                                    if import_type == "cod_payment":
                                        existing_val = (
                                            existing.get("cod_amount")
                                            or existing.get("amount")
                                            or 0.0
                                        )
                                    else:
                                        existing_val = (
                                            existing.get("amount")
                                            or existing.get("cod_amount")
                                            or 0.0
                                        )
                                else:
                                    existing_val = existing.get(field)
                                if field == "weight":
                                    equal = _weight_equal(existing_val, new_val)
                                    row_imported_weight = str(new_val)
                                elif field == "payment_mode":
                                    equal = _payment_mode_equal(existing_val, new_val)
                                    row_imported_payment_mode = str(new_val)
                                else:  # amount
                                    equal = _amount_equal(existing_val, new_val)
                                    try:
                                        row_imported_amount = float(new_val)
                                    except (TypeError, ValueError):
                                        row_imported_amount = None
                                if not equal:
                                    mismatches.append({
                                        "field":    field,
                                        "existing": existing_val,
                                        "imported": new_val,
                                    })
                                    # DO NOT override — user requirement.
                                    continue

                            # Special mapping: courier_hint drives courier_name.
                            if field == "courier_hint":
                                applied["courier_name"] = new_val
                                continue
                            if field == "address":
                                applied["address_line1"] = new_val
                                continue
                            if field == "items":
                                # Items in DB is a list; store as single-element list
                                # when incoming is a plain string.
                                applied["items"] = [new_val] if isinstance(new_val, str) and new_val else new_val
                                continue
                            applied[field] = new_val

                        # ─── Phase F6.2 — Type-specific import stamps ───
                        # These fields ADD to (never replace) the merchant's
                        # original booking data. The Shipment Details screen
                        # reads them to render the "Import" panel.
                        if import_type == "booking":
                            applied["imported_booking_at"] = now
                            applied["last_import_at"] = now
                            applied["last_import_type"] = "booking"
                            if row_imported_weight:
                                applied["imported_post_office_weight"] = row_imported_weight
                            if row_imported_payment_mode:
                                applied["imported_courier_payment_mode"] = row_imported_payment_mode
                            if row_imported_amount is not None:
                                applied["imported_booked_cod_amount"] = row_imported_amount
                            # Persist current-run validation alerts on the shipment
                            # so the Details screen can display badges. Cross-verify
                            # mismatches from THIS row become the shipment's live
                            # alerts (overwrites previous run's alerts by design —
                            # newest import is the source of truth).
                            applied["import_validation_alerts"] = [
                                {
                                    "field":    mm["field"],
                                    "existing": mm["existing"],
                                    "imported": mm["imported"],
                                    "at":       now,
                                }
                                for mm in mismatches
                            ]

                        # Type-specific post-processing.
                        if import_type == "delivery":
                            # Phase F8.0 — routed through the SHARED
                            # Shipment Update Engine (see
                            # apply_last_event_engine above). The SMS
                            # ingest pipeline (routers/courier_sync.py)
                            # calls the SAME function, so import + SMS
                            # can never drift on business rules:
                            #   • last_event_category badge
                            #   • F7.7 status routing (Delivered /
                            #     In Transit / Out for Delivery)
                            #   • Return Review flag (no auto-move)
                            #   • delivered_at stamp
                            #   • confirmation_status="confirmed"
                            raw_ev = applied.get("last_event") or ""
                            apply_last_event_engine(
                                existing, applied,
                                raw_event=raw_ev,
                                now=now,
                                source="import",
                            )

                            applied["delivery_source"] = "imported"
                            applied["last_import_at"] = now
                            applied["last_import_type"] = "delivery"

                        if import_type == "cod_payment":
                            # Stamp the courier collection metadata on the shipment
                            # for audit trail. These live on top of any existing
                            # values; if user re-uploads later the newest wins.
                            if not applied.get("cod_payment_date"):
                                applied["cod_payment_date"] = now
                            applied["cod_payment_status"] = "received"
                            applied["last_import_at"] = now
                            applied["last_import_type"] = "cod_payment"

                            # Phase F10.2 — Payment discrepancy detection.
                            # Explicitly compare the courier-remitted amount
                            # (`cod_collected_amount` uploaded in this row)
                            # against the shipment's stored `cod_amount`
                            # (what the courier was supposed to collect).
                            # If they differ, register a mismatch so:
                            #   • row_status flips to matched_mismatch below
                            #   • the batch tally counts it under mismatches
                            #   • the Shipment Details "Payment discrepancy"
                            #     alert surfaces it in-app
                            # cod_collected_amount isn't in CROSS_VERIFY_FIELDS
                            # so we handle the compare here as a first-class
                            # concern for cod_payment imports only.
                            uploaded_cod = applied.get("cod_collected_amount")
                            stored_cod   = existing.get("cod_amount")
                            if (
                                uploaded_cod is not None
                                and stored_cod is not None
                                and not _amount_equal(stored_cod, uploaded_cod)
                                # Don't double-count if the cross-verify
                                # loop already flagged an amount mismatch
                                # for this row.
                                and not any(
                                    m.get("field") in ("amount", "cod_amount")
                                    for m in mismatches
                                )
                            ):
                                mismatches.append({
                                    "field":    "cod_amount",
                                    "existing": stored_cod,
                                    "imported": uploaded_cod,
                                })

                            # Phase F10.2 — Project current-run mismatches
                            # onto the shipment doc so the Shipment
                            # Details screen surfaces the "Payment
                            # discrepancy" alert. Mirrors the booking-
                            # branch behaviour (lines 1134-1142) that
                            # was previously scoped only to booking
                            # imports. Overwrites any prior alerts by
                            # design — the newest cod_payment run is
                            # the source of truth for the alert state.
                            if mismatches:
                                applied["import_validation_alerts"] = [
                                    {
                                        "field":    mm["field"],
                                        "existing": mm["existing"],
                                        "imported": mm["imported"],
                                        "at":       now,
                                    }
                                    for mm in mismatches
                                ]

                        if applied:
                            row_status = "matched_mismatch" if mismatches else "matched_updated"
                            updates_to_write.append((shipment_id, applied))
                        else:
                            row_status = "matched_mismatch" if mismatches else "matched_no_change"
            except Exception as e:
                row_status = "error"
                row_error = str(e)[:240]

            row_entry: Dict[str, Any] = {
                "row_index":   idx + data_start_row,   # 1-based row in the ORIGINAL file
                "tracking_id": tracking,
                "status":      row_status,
                "shipment_id": shipment_id,
                "applied":     applied,
                "mismatches":  mismatches,
            }
            # Phase F7.6 — per-row booking_date outcome + batch tallies.
            if booking_date_row_status:
                row_entry["booking_date_status"] = booking_date_row_status
                if booking_date_row_status == "updated":
                    batch_doc["booking_date_updated"] += 1
                elif booking_date_row_status == "skipped":
                    batch_doc["booking_date_skipped"] += 1
                elif booking_date_row_status == "invalid":
                    batch_doc["booking_date_invalid"] += 1
            if row_error:
                row_entry["error"] = row_error
            batch_doc["rows"].append(row_entry)

            if row_status == "matched_updated":
                batch_doc["matched_updated"] += 1
            elif row_status == "matched_mismatch":
                batch_doc["matched_mismatch"] += 1
            elif row_status == "matched_no_change":
                batch_doc["matched_no_change"] += 1
            elif row_status == "unmatched":
                batch_doc["unmatched"] += 1
            elif row_status == "junk_skipped":
                batch_doc["junk_skipped"] += 1
            elif row_status == "error":
                batch_doc["errors"] += 1

        # ────────────── Write updates in one bulk op ──────────────
        wrote = 0
        settled_shipment_ids: List[str] = []
        settled_total_amount: float = 0.0
        if updates_to_write:
            from pymongo import UpdateOne  # local import to avoid top-level cost
            ops = []
            for sid, upd in updates_to_write:
                # Stamp modified metadata so the Shipments UI's "Modified"
                # bucket picks these up if the merchant wants to review.
                upd_final = dict(upd)
                upd_final["is_modified"] = True
                upd_final["modified_at"] = now
                # Phase F6.3 — always link the shipment to THIS import
                # batch so the Filter panel's "Import Batch" chip can
                # drill down. Stored as an array (a shipment can be
                # touched by multiple imports of different types over
                # its lifetime).
                # Using $addToSet via UpdateOne needs $set + $addToSet
                # split — see the ops build below.
                # For delivery, also stamp the pipeline delivered_at when
                # status hits Delivered. Existing helper checks in
                # shipments_write.py will not fire here (bulk write path),
                # so we mirror the fields manually.
                if upd_final.get("status") == "Delivered" and not upd_final.get("delivered_at"):
                    upd_final["delivered_at"] = now
                ops.append(UpdateOne(
                    {"id": sid, "user_id": current_user["id"]},
                    {
                        "$set": upd_final,
                        "$addToSet": {"import_batch_ids": batch_id},
                    },
                ))
                if import_type == "cod_payment":
                    settled_shipment_ids.append(sid)
                    try:
                        settled_total_amount += float(upd_final.get("cod_collected_amount") or 0.0)
                    except (TypeError, ValueError):
                        pass
            if ops:
                res = await db.shipments.bulk_write(ops, ordered=False)
                wrote = int(res.modified_count or 0)

        # Phase F6.3 — create the PaymentBatch AFTER writes so we can
        # attach the actual list of settled shipments + real total.
        pb_created: Optional[Dict[str, Any]] = None
        if payment_batch_payload and import_type == "cod_payment":
            from routers.payment_batches import create_payment_batch as _create_pb
            pb_created = await _create_pb(
                db, current_user["id"], payment_batch_payload,
                shipment_ids=settled_shipment_ids,
                total_amount=settled_total_amount,
                import_batch_id=batch_id,
            )
            # Stamp payment_batch_id on each settled shipment so the
            # Payment Batch filter chip can find them without a join.
            if settled_shipment_ids and pb_created:
                await db.shipments.update_many(
                    {"user_id": current_user["id"], "id": {"$in": settled_shipment_ids}},
                    {"$set": {"payment_batch_id": pb_created["id"]}},
                )
            # Record on the import batch record for cross-lookup.
            batch_doc["payment_batch_id"] = pb_created["id"]

        # Persist the batch record for history / audit.
        try:
            await db.shipment_import_batches.insert_one(batch_doc)
        except Exception:
            _logger.exception("failed to persist shipment_import batch")

        _logger.info(
            "shipment-import: user=%s type=%s file=%s rows=%d matched=%d mismatch=%d unmatched=%d",
            current_user["id"], import_type, file.filename or "",
            len(rows), batch_doc["matched_updated"], batch_doc["matched_mismatch"],
            batch_doc["unmatched"],
        )

        return {
            "ok":                True,
            "batch_id":          batch_id,
            "import_type":       import_type,
            "filename":          batch_doc["filename"],
            "format":            fmt,
            "total_rows":        batch_doc["total_rows"],
            "matched_updated":   batch_doc["matched_updated"],
            "matched_mismatch":  batch_doc["matched_mismatch"],
            "matched_no_change": batch_doc["matched_no_change"],
            "unmatched":         batch_doc["unmatched"],
            "errors":            batch_doc["errors"],
            # Phase F11.A — rows dropped as spreadsheet junk (Total /
            # RS / N-A / etc). Counted separately so the operator can
            # tell the difference between "10 rows didn't match my DB"
            # (real problem) and "10 rows were summary noise".
            "junk_skipped":      batch_doc["junk_skipped"],
            "db_modified":       wrote,
            # Phase F7.6 — Delivery-import booking_date fill counters.
            # Present on ALL commit responses (zero for non-delivery
            # imports) so the frontend can conditionally render the
            # summary section without a defensive branch per type.
            "booking_date_updated": batch_doc["booking_date_updated"],
            "booking_date_skipped": batch_doc["booking_date_skipped"],
            "booking_date_invalid": batch_doc["booking_date_invalid"],
            # Phase F6.3 — surface the payment batch if one was created.
            "payment_batch":     pb_created,
            "duplicate_warning": duplicate_batch if (pb_created and duplicate_batch) else None,
        }

    # ── History / Audit ───────────────────────────────────────────

    @shipment_import_router.get("/shipments/import/batches")
    async def list_batches(
        current_user: Dict[str, Any] = Depends(_get_current_user),
        limit: int = Query(50, ge=1, le=200),
    ):
        cursor = db.shipment_import_batches.find(
            {"user_id": current_user["id"]},
            {"_id": 0, "rows": 0},   # summary view only
        ).sort("created_at", -1).limit(limit)
        return {"batches": [doc async for doc in cursor]}

    @shipment_import_router.get("/shipments/import/batches/{batch_id}")
    async def get_batch(
        batch_id: str,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        doc = await db.shipment_import_batches.find_one(
            {"id": batch_id, "user_id": current_user["id"]},
            {"_id": 0},
        )
        if not doc:
            raise HTTPException(status_code=404, detail="Batch not found")
        return doc

    # ── Delete a batch history record ─────────────────────────────
    #
    # This ONLY removes the audit trail entry for a bulk import
    # (Booking / Delivery Status / COD Payment). Shipment records
    # that were already updated by the import remain updated — this
    # endpoint does NOT revert any changes applied to shipments or
    # to related payment batches. That's a conscious trade-off: a
    # full undo would be prohibitively expensive to track.
    @shipment_import_router.delete("/shipments/import/batches/{batch_id}")
    async def delete_batch(
        batch_id: str,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        res = await db.shipment_import_batches.delete_one(
            {"id": batch_id, "user_id": current_user["id"]},
        )
        if res.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Batch not found")
        return {"ok": True, "deleted": batch_id}

    @shipment_import_router.get("/shipments/import/batches/{batch_id}/mismatches.csv")
    async def download_mismatches(
        batch_id: str,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        doc = await db.shipment_import_batches.find_one(
            {"id": batch_id, "user_id": current_user["id"]},
            {"_id": 0},
        )
        if not doc:
            raise HTTPException(status_code=404, detail="Batch not found")
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "row_index", "tracking_id", "status", "field",
            "existing_value", "imported_value", "shipment_id", "error",
        ])
        for r in doc.get("rows", []):
            status = r.get("status")
            # Include unmatched, mismatch, and error rows in the report so
            # the merchant sees the full downloadable audit.
            if status == "matched_mismatch":
                for mm in r.get("mismatches", []):
                    w.writerow([
                        r.get("row_index"), r.get("tracking_id"), status,
                        mm.get("field"),
                        _csv_val(mm.get("existing")),
                        _csv_val(mm.get("imported")),
                        r.get("shipment_id") or "", r.get("error") or "",
                    ])
            elif status in ("unmatched", "error"):
                w.writerow([
                    r.get("row_index"), r.get("tracking_id"), status,
                    "", "", "", r.get("shipment_id") or "", r.get("error") or "",
                ])
        headers = {
            "Content-Disposition": f'attachment; filename="mismatches_{batch_id[:8]}.csv"',
        }
        return Response(content=buf.getvalue(), media_type="text/csv", headers=headers)

    # ── Article Number Mismatch (Phase F11.A) ────────────────────
    #
    # Aggregated view of tracking numbers that appeared in bulk
    # uploads (booking / delivery / cod_payment) but did NOT match
    # any shipment in the merchant's DB. Junk noise (Total / RS /
    # etc) is filtered out at import time, so what surfaces here is
    # the real backlog — courier remit rows for shipments the
    # merchant hasn't booked yet in the app.
    #
    # Response shape:
    #   {
    #     "items": [
    #       {
    #         "tracking_id":    "AB123456789IN",
    #         "occurrence":     3,
    #         "last_import_type": "cod_payment",
    #         "last_seen":      "...ISO...",
    #         "batches":        [{id, filename, import_type, created_at}]
    #       }
    #     ],
    #     "total": 12
    #   }
    @shipment_import_router.get("/shipments/import/article-mismatches")
    async def article_mismatches(
        current_user: Dict[str, Any] = Depends(_get_current_user),
        import_type: Optional[str] = Query(None, description="booking|delivery|cod_payment"),
        limit: int = Query(500, ge=1, le=5000),
    ):
        if import_type is not None and import_type not in IMPORT_TYPES:
            raise HTTPException(
                status_code=400,
                detail=f"import_type must be one of {sorted(IMPORT_TYPES)}",
            )
        # Aggregate across the last N batches (cap for perf) so old
        # historical batches don't blow this up on large tenants.
        query: Dict[str, Any] = {"user_id": current_user["id"]}
        if import_type:
            query["import_type"] = import_type
        cursor = db.shipment_import_batches.find(
            query, {"_id": 0}
        ).sort("created_at", -1).limit(200)

        agg: Dict[str, Dict[str, Any]] = {}
        async for batch in cursor:
            b_meta = {
                "id":          batch.get("id"),
                "filename":    batch.get("filename") or "",
                "import_type": batch.get("import_type"),
                "created_at":  batch.get("created_at"),
            }
            for row in (batch.get("rows") or []):
                if row.get("status") != "unmatched":
                    continue
                tid_raw = row.get("tracking_id") or ""
                tid = str(tid_raw).strip()
                if not tid or _is_junk_tracking_value(tid):
                    continue
                key = tid.lower()
                bucket = agg.get(key)
                if not bucket:
                    bucket = {
                        "tracking_id":      tid,
                        "occurrence":       0,
                        "last_import_type": b_meta["import_type"],
                        "last_seen":        b_meta["created_at"],
                        "batches":          [],
                    }
                    agg[key] = bucket
                bucket["occurrence"] += 1
                # First iteration = newest batch (sorted DESC), so
                # last_import_type / last_seen stay pinned to it.
                # Append batch context (dedup by id).
                if not any(x["id"] == b_meta["id"] for x in bucket["batches"]):
                    bucket["batches"].append(b_meta)

        items = sorted(
            agg.values(),
            key=lambda x: (x.get("last_seen") or "", x.get("occurrence") or 0),
            reverse=True,
        )[:limit]
        return {"items": items, "total": len(agg)}

    _logger.info("shipment_import router mounted (10 endpoints)")


def _csv_val(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)
