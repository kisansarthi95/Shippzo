"""
File Import — Phase F1 (CSV + XLSX bulk-import to pending_orders).

Mirrors the Smart Paste / Google Sheet ingestion path, but the source
is a user-uploaded `.csv` or `.xlsx` file. Once a row is parsed and
validated, it lands in the same `pending_orders` Mongo collection
(source="file") so the Orders tab + Ship-this-order pipeline reuse
the existing UI + business rules unchanged.

Endpoints:
  POST   /api/orders/import/preview     parse-only, return columns + first N sample rows
  POST   /api/orders/import/commit      parse + map + insert pending_orders
  GET    /api/me/file-import-mapping    saved default column → schema mapping
  PUT    /api/me/file-import-mapping    save / update the default mapping

Pattern: late-binding `init()` — same as routers/notifications.py.
"""
from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel


file_import_router = APIRouter(prefix="/api", tags=["file-import"])

# ────────────────────────────────────────────────────────────────────
# Schema fields the user can map columns to.
#
# Phase F1.1 (2026-05-08): The system stores a delivery address in
# address_line1 / address_line2 internally, but the IMPORT UX exposes
# a single virtual "address" field — many CSV/Excel exports already
# have separate "Address Line 1" / "Address Line 2" columns and asking
# the user to map both to two distinct schema fields invites
# truncation / line2 falling off when the user only maps line1.
# Instead, mapping MULTIPLE columns to "address" auto-merges them
# (in mapping-iteration order) with a single-space separator and
# writes the result to address_line1.
#
# Phase F1.2 (2026-05-08): Mirrored every field in the Add-Shipment
# form so a user can map ANY column from their CSV/Excel — box size,
# parcel dimensions (L/W/H), category, plus all customer + payment
# fields. List is ordered by frequency-of-use so the most common
# fields (Customer Name → Phone → Address → Amount → Token) sit at
# the top of the field-picker modal where users actually look.
# ────────────────────────────────────────────────────────────────────
SCHEMA_FIELDS: List[str] = [
    # ── Identity (most commonly mapped) ──
    "customer_name",
    "customer_phone",
    "customer_alt_phone",
    "customer_email",
    "customer_gstin",
    # ── Delivery address ──
    "address",                  # virtual; persisted to address_line1
    "city",
    "state",
    "pincode",
    # ── Payment ──
    "amount",
    "token_amount",
    "payment_mode",
    # ── Items / parcel ──
    "items",
    "category",
    "weight",
    "box_dimensions",           # free-form, e.g. "10x8x4"
    "box_length",               # numeric (cm)
    "box_width",                # numeric (cm)
    "box_height",               # numeric (cm)
    # ── Misc ──
    "courier_hint",
    "order_id",
    "notes",
]
NUMERIC_FIELDS = {
    "amount", "token_amount",
    "box_length", "box_width", "box_height",
    # weight stays as string ("250g", "0.5 kg" etc.) — matches the
    # existing PendingOrder.weight: str schema.
}
PAYMENT_MODE_NORMALISE = {
    "cod": "COD", "c": "COD", "cash on delivery": "COD",
    "paid": "PAID", "p": "PAID", "prepaid": "PAID",
    "online": "PAID", "upi": "PAID",
}

MAX_SAMPLE_ROWS = 10
MAX_FILE_BYTES  = 10 * 1024 * 1024     # 10 MB hard limit
MAX_IMPORT_ROWS = 5000                 # one upload at a time


class FileImportMappingPayload(BaseModel):
    mapping: Dict[str, str]            # { csv_column_header: schema_field }


# ════════════════════════════════════════════════════════════════════
#                              Parsers
# ════════════════════════════════════════════════════════════════════

def _parse_csv(blob: bytes) -> tuple[List[str], List[Dict[str, str]]]:
    # CSV may use BOM (Excel-saved csv) — decode-detect.
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = blob.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(status_code=400, detail="Could not decode file")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File is empty")
    rows: List[Dict[str, str]] = []
    for line in reader:
        # pad / trim to header length
        if len(line) < len(header):
            line = list(line) + [""] * (len(header) - len(line))
        elif len(line) > len(header):
            line = line[: len(header)]
        rows.append({h.strip(): (line[i] or "").strip() for i, h in enumerate(header)})
    return [h.strip() for h in header], rows


def _parse_xlsx(blob: bytes) -> tuple[List[str], List[Dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="XLSX parser (openpyxl) not installed on the server",
        )
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read XLSX: {e}")
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Workbook has no sheets")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Sheet is empty")
    header = [str(c).strip() if c is not None else "" for c in header_raw]
    rows: List[Dict[str, str]] = []
    for line in rows_iter:
        line = list(line) + [None] * (len(header) - len(line))
        line = line[: len(header)]
        if all(c is None or str(c).strip() == "" for c in line):
            continue   # skip blank rows
        rows.append({
            h: ("" if line[i] is None else str(line[i]).strip())
            for i, h in enumerate(header)
        })
    return header, rows


def _parse_upload(file: UploadFile, blob: bytes) -> tuple[str, List[str], List[Dict[str, str]]]:
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        cols, rows = _parse_xlsx(blob)
        return "xlsx", cols, rows
    if name.endswith(".csv") or (file.content_type or "").startswith("text/csv"):
        cols, rows = _parse_csv(blob)
        return "csv", cols, rows
    raise HTTPException(
        status_code=400,
        detail="Only .csv and .xlsx files are supported (got "
               f"{file.filename or 'unknown'})",
    )


def _normalise_value(field: str, raw: str) -> Any:
    raw = (raw or "").strip()
    if field in NUMERIC_FIELDS:
        if not raw:
            return 0.0
        try:
            return float(raw.replace(",", "").replace("₹", "").strip())
        except ValueError:
            return 0.0
    if field == "payment_mode":
        return PAYMENT_MODE_NORMALISE.get(raw.lower(), "COD" if not raw else raw.upper())
    if field == "pincode":
        # keep as string but strip non-digits
        return "".join(ch for ch in raw if ch.isdigit())
    return raw


# ════════════════════════════════════════════════════════════════════
#                              Routes
# ════════════════════════════════════════════════════════════════════

def init() -> None:
    """Register routes after server.py finishes initialising."""
    import logging
    _logger = logging.getLogger("routers.file_import")
    from server import (  # noqa: WPS433 — late-binding to avoid circular import
        db,
        get_current_user as _get_current_user,
        generate_master_order_id,
    )

    # ── Saved default mapping (per user) ──────────────────────────

    @file_import_router.get("/me/file-import-mapping")
    async def get_file_import_mapping(
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        doc = await db.users.find_one(
            {"id": current_user["id"]}, {"_id": 0, "file_import_mapping": 1},
        ) or {}
        return {
            "mapping":     doc.get("file_import_mapping") or {},
            "schema_fields": SCHEMA_FIELDS,
        }

    @file_import_router.put("/me/file-import-mapping")
    async def put_file_import_mapping(
        payload: FileImportMappingPayload,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        # Sanity-check the mapping values are known schema fields.
        bad = [v for v in payload.mapping.values() if v and v not in SCHEMA_FIELDS]
        if bad:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown schema fields: {sorted(set(bad))}",
            )
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": {"file_import_mapping": payload.mapping}},
        )
        return {"ok": True, "mapping": payload.mapping}

    # ── Preview (no DB write) ─────────────────────────────────────

    @file_import_router.post("/orders/import/preview")
    async def import_preview(
        file: UploadFile = File(...),
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        blob = await file.read()
        if len(blob) > MAX_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large (limit { MAX_FILE_BYTES // (1024 * 1024) } MB)",
            )
        fmt, columns, rows = _parse_upload(file, blob)
        # Auto-suggest mapping using the saved default + naive header match.
        saved = (
            await db.users.find_one(
                {"id": current_user["id"]}, {"_id": 0, "file_import_mapping": 1},
            ) or {}
        ).get("file_import_mapping") or {}
        # Heuristic header→field aliases (lowercase, _-collapsed key).
        # Phase F1.1 — extra entries route any "address line *" / "addr"
        # variant to the single virtual `address` field so users don't
        # have to know about line1/line2 plumbing.
        ALIASES = {
            "address":         "address",
            "addr":            "address",
            "address_line":    "address",
            "address_line_1":  "address",
            "address_line_2":  "address",
            "address_1":       "address",
            "address_2":       "address",
            "addressline1":    "address",
            "addressline2":    "address",
            "full_address":    "address",
            "delivery_address":"address",
            "name":            "customer_name",
            "customer":        "customer_name",
            "phone":           "customer_phone",
            "mobile":          "customer_phone",
            "contact":         "customer_phone",
            "alt_phone":       "customer_alt_phone",
            "email":           "customer_email",
            "gst":             "customer_gstin",
            "gstin":           "customer_gstin",
            "pin":             "pincode",
            "pin_code":        "pincode",
            "zip":             "pincode",
            "amt":             "amount",
            "total":           "amount",
            "order_amount":    "amount",
            "order_value":     "amount",
            "token":           "token_amount",
            "advance":         "token_amount",
            "advance_amount":  "token_amount",
            "token_amt":       "token_amount",
            "mode":            "payment_mode",
            "payment":         "payment_mode",
            "courier":         "courier_hint",
            "logistics":       "courier_hint",
            "wt":              "weight",
            "parcel_weight":   "weight",
            "package_weight":  "weight",
            "remarks":         "notes",
            "comment":         "notes",
            "comments":        "notes",
            "shipment_notes":  "notes",
            # Phase F1.2 — box / parcel
            "box":             "box_dimensions",
            "box_size":        "box_dimensions",
            "dimensions":      "box_dimensions",
            "lwh":             "box_dimensions",
            "size":            "box_dimensions",
            "length":          "box_length",
            "l":               "box_length",
            "breadth":         "box_width",
            "width":           "box_width",
            "w":               "box_width",
            "height":          "box_height",
            "h":               "box_height",
            "category":        "category",
            "cat":             "category",
            "item_category":   "category",
            "item":            "items",
            "product":         "items",
            "products":        "items",
        }
        suggested: Dict[str, str] = {}
        for c in columns:
            if c in saved:
                suggested[c] = saved[c]
                continue
            cl = c.lower().strip().replace(" ", "_").replace("-", "_")
            if cl in SCHEMA_FIELDS:
                suggested[c] = cl
            elif cl in ALIASES:
                suggested[c] = ALIASES[cl]
        return {
            "format":         fmt,
            "filename":       file.filename or "",
            "columns":        columns,
            "sample_rows":    rows[:MAX_SAMPLE_ROWS],
            "total_rows":     len(rows),
            "schema_fields":  SCHEMA_FIELDS,
            "suggested":      suggested,
        }

    # ── Commit (validate + insert) ────────────────────────────────

    @file_import_router.post("/orders/import/commit")
    async def import_commit(
        file: UploadFile = File(...),
        mapping: str = Form(..., description="JSON string {col_header: schema_field}"),
        save_default: bool = Form(False),
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        try:
            mapping_obj: Dict[str, str] = json.loads(mapping)
            if not isinstance(mapping_obj, dict):
                raise ValueError("mapping must be a JSON object")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid mapping JSON: {e}")

        blob = await file.read()
        if len(blob) > MAX_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large (limit { MAX_FILE_BYTES // (1024 * 1024) } MB)",
            )
        fmt, columns, rows = _parse_upload(file, blob)
        if len(rows) > MAX_IMPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"Too many rows ({len(rows)}). Max {MAX_IMPORT_ROWS} per "
                    "upload — please split the file."
                ),
            )

        # Reverse-lookup: schema_field → list[file_column].
        # Phase F1.1 — multiple columns can map to the same field
        # (specifically the virtual "address" field, which auto-merges
        # Address Line 1 + Address Line 2 from the source file). For
        # all other fields the LAST mapping wins (sane fallback if a
        # user accidentally maps two columns to e.g. `customer_name`).
        field_to_cols: Dict[str, List[str]] = {}
        for col, field in mapping_obj.items():
            if not field:
                continue
            if field not in SCHEMA_FIELDS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unknown schema field in mapping: {field}",
                )
            field_to_cols.setdefault(field, []).append(col)

        if (not field_to_cols.get("customer_name")
                and not field_to_cols.get("customer_phone")):
            raise HTTPException(
                status_code=400,
                detail=(
                    "At minimum, map `customer_name` or `customer_phone` "
                    "before importing."
                ),
            )

        if save_default:
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"file_import_mapping": mapping_obj}},
            )

        # Build pending_orders rows.
        now = datetime.now(timezone.utc).isoformat()
        imported = 0
        skipped = 0
        errors:  List[str] = []
        batch:   List[Dict[str, Any]] = []
        filename = file.filename or "import"
        # Pre-allocate master_order_ids in bulk to avoid per-row counter
        # round-trips.
        master_oids = [
            await generate_master_order_id() for _ in range(len(rows))
        ]
        for idx, row in enumerate(rows):
            try:
                doc: Dict[str, Any] = {
                    "id":                str(uuid.uuid4()),
                    "user_id":           current_user["id"],
                    "source":            "file",
                    "status":            "pending",
                    "master_order_id":   master_oids[idx],
                    "order_id":          "",
                    "created_at":        now,
                    # Phase-F1 — extra metadata for the Orders UI badge.
                    "source_meta": {
                        "filename":     filename,
                        "format":       fmt,
                        "imported_at":  now,
                        "row_index":    idx + 2,   # +2 = 1 (header) + 1-based
                    },
                }
                # Initialise all schema fields to defaults.
                # Note: the virtual "address" field doesn't exist in
                # the PendingOrder schema — we materialise it into
                # address_line1 below. Storage stays line1 + line2 so
                # downstream Sheet append + label rendering continue
                # working unchanged.
                for field in SCHEMA_FIELDS:
                    if field == "address":
                        continue
                    doc[field] = 0.0 if field in NUMERIC_FIELDS else (
                        "COD" if field == "payment_mode" else ""
                    )
                doc["address_line1"] = ""
                doc["address_line2"] = ""

                # Fill from mapped columns. For "address", concatenate
                # all mapped columns in mapping-iteration order with a
                # single-space separator (line1 + " " + line2 etc.) and
                # write to address_line1.
                for field, cols in field_to_cols.items():
                    if field == "address":
                        parts = [
                            (row.get(c, "") or "").strip()
                            for c in cols
                        ]
                        doc["address_line1"] = " ".join(p for p in parts if p)
                        continue
                    raw = row.get(cols[-1], "")  # last-mapped wins
                    doc[field] = _normalise_value(field, raw)
                doc["order_id"] = doc.get("order_id") or doc["master_order_id"]
                # Skip blank rows (no name AND no phone)
                if not doc["customer_name"] and not doc["customer_phone"]:
                    skipped += 1
                    continue
                batch.append(doc)
                imported += 1
            except Exception as e:
                skipped += 1
                errors.append(f"row {idx + 2}: {e}")

        if batch:
            try:
                await db.pending_orders.insert_many(batch, ordered=False)
            except Exception as e:
                _logger.exception("pending_orders insert_many failed")
                raise HTTPException(status_code=500, detail=f"DB write failed: {e}")

        _logger.info(
            "file-import: user=%s file=%s fmt=%s imported=%d skipped=%d",
            current_user["id"], filename, fmt, imported, skipped,
        )
        return {
            "ok":         True,
            "imported":   imported,
            "skipped":    skipped,
            "total":      len(rows),
            "errors":     errors[:20],   # cap so the response stays small
            "filename":   filename,
            "format":     fmt,
        }

    _logger.info("file_import router mounted (4 endpoints)")
