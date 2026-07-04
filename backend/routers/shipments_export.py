"""
Shipments — CSV + XLSX export endpoints — Phase F4.6 refactor.

Moved out of `routers/shipments_read.py` so a single router file
doesn't grow into a grab-bag of everything shipment-related. Nothing
about the API contract changed — the four endpoints below still live
at exactly the same paths:

    GET  /api/shipments/export/csv          → all shipments, CSV
    POST /api/shipments/export/csv          → filtered id list, CSV
    GET  /api/shipments/export/xlsx         → all shipments, XLSX
    POST /api/shipments/export/xlsx         → filtered id list, XLSX

Registration order matters: this router MUST be included BEFORE any
router that owns a `/shipments/{shipment_id}` catch-all GET, otherwise
FastAPI will route `/shipments/export/csv` into the catch-all.
Server.py handles the ordering; see the corresponding `include_router`
block.

Why bother with a separate file?
  1. `shipments_read.py` was pushing 800+ lines with three concerns
     (list/stats/lookup + row-builder + Excel encoding). Isolating the
     row-builder + openpyxl imports keeps the read path lean.
  2. Openpyxl is a fairly heavy transitive import — keeping it in its
     own module makes the cost obvious at code-review time.
  3. Future export formats (XLSX-with-images for label previews, PDF
     shipment manifest, etc.) live naturally in this file.
"""
from __future__ import annotations

import csv
import io
from io import BytesIO
from typing import Any, Dict, List, Optional

from pydantic import BaseModel

from fastapi import APIRouter, Depends
from fastapi.responses import PlainTextResponse, Response


shipments_export_router = APIRouter(prefix="/api", tags=["shipments-export"])


# ─── request body ────────────────────────────────────────────────────
# The client posts the EXACT id list currently visible on screen so
# the server doesn't have to duplicate the compound status + date
# range filter logic. When missing/empty, all owned shipments are
# exported.
class _ExportCsvBody(BaseModel):
    ids: Optional[List[str]] = None


# `_build_shipment_rows_for_user`, `_build_csv_for_user` and
# `_build_xlsx_bytes` are constructed inside init() so they can close
# over the late-bound `db` reference (Mongo client isn't ready at
# import time).
def init() -> None:
    """Register export routes after server.py has initialised db + auth."""
    import logging
    _logger = logging.getLogger("routers.shipments_export")
    from server import (  # noqa: WPS433 — intentional late import
        db,
        get_current_user as _get_current_user,
    )

    # ─── row builder — SHARED by CSV + XLSX ──────────────────────────
    async def _build_shipment_rows_for_user(
        user_id: str,
        ids: Optional[List[str]] = None,
    ) -> tuple[List[str], List[List[Any]]]:
        """Return (COLUMNS, rows) for the current user's shipments.

        Single source of truth for column order and value coercion —
        the CSV writer, the openpyxl writer, and any future export
        format share the exact same schema.
        """
        mongo_q: Dict[str, Any] = {"user_id": user_id}
        if ids:
            mongo_q["id"] = {"$in": list(ids)[:10_000]}
        docs = await db.shipments.find(mongo_q, {"_id": 0}).sort(
            "created_at", -1
        ).to_list(10_000)
        COLUMNS = [
            "Shipment ID",          "Tracking ID",          "Master Order ID",
            "Order ID",             "AWB Number",
            "Status",               "Payment Status",
            "Customer Name",        "Phone",                "Alt Phone",
            "Email",
            "Address Line 1",       "Address Line 2",       "Landmark",
            "City",                 "State",                "Pincode",        "Country",
            "Courier",              "Courier Service",      "Courier Tracking URL",
            "Items",                "Item Description",     "Quantity",
            "Weight",               "Box Dimensions",
            "Payment Mode",         "Payment Type",         "Amount",         "Token Amount",
            "COD Balance",
            "Discount",             "Tax",                  "Shipping Charges",
            "Notes",                "Shipment Notes",       "Internal Notes",
            "Source",
            "Created At",           "Updated At",
            "Shipped At",           "Out For Delivery At",  "Delivered At",
            "Cancelled At",         "Returned At",
        ]
        rows: List[List[Any]] = []
        for d in docs:
            items = d.get("items") or []
            items_str = "; ".join(items) if items else d.get("item_description", "")
            amount = float(d.get("amount") or 0)
            token  = float(d.get("token_amount") or d.get("token") or 0)
            cod_balance = ""
            ptype = (d.get("payment_type") or d.get("payment_mode") or "").upper()
            if ptype == "COD" or ptype == "COD/PARTIAL":
                if "cod_amount" in d and d.get("cod_amount") is not None:
                    cod_balance = f"{float(d['cod_amount'] or 0):.2f}"
                else:
                    cod_balance = f"{max(amount - token, 0):.2f}"
            rows.append([
                d.get("id", ""),
                d.get("tracking_id", ""),
                d.get("master_order_id", ""),
                d.get("order_id", ""),
                d.get("awb_number", "") or d.get("awb", ""),
                d.get("status", ""),
                d.get("payment_status", ""),
                d.get("customer_name", ""),
                d.get("customer_phone", ""),
                d.get("customer_alt_phone", "") or d.get("alt_phone", ""),
                d.get("customer_email", ""),
                d.get("address_line1", ""),
                d.get("address_line2", ""),
                d.get("landmark", ""),
                d.get("city", ""),
                d.get("state", ""),
                d.get("pincode", ""),
                d.get("country", "") or "India",
                d.get("courier_name", "") or d.get("courier", ""),
                d.get("courier_service", "") or d.get("service_type", ""),
                d.get("courier_tracking_url", "") or d.get("tracking_url", ""),
                items_str,
                d.get("item_description", ""),
                d.get("quantity", "") or d.get("qty", ""),
                d.get("weight", ""),
                d.get("box_dimensions", "") or d.get("dimensions", ""),
                d.get("payment_mode", ""),
                d.get("payment_type", ""),
                f"{amount:.2f}" if amount else "",
                f"{token:.2f}"  if token  else "",
                cod_balance,
                d.get("discount", ""),
                d.get("tax", ""),
                d.get("shipping_charges", "") or d.get("shipping", ""),
                d.get("notes", ""),
                d.get("shipment_notes", ""),
                d.get("internal_notes", ""),
                d.get("source", "") or d.get("created_via", ""),
                d.get("created_at", ""),
                d.get("updated_at", ""),
                d.get("shipped_at", ""),
                d.get("out_for_delivery_at", ""),
                d.get("delivered_at", ""),
                d.get("cancelled_at", ""),
                d.get("returned_at", ""),
            ])
        return COLUMNS, rows

    # ─── CSV builder ────────────────────────────────────────────────
    async def _build_csv_for_user(
        user_id: str, ids: Optional[List[str]] = None,
    ) -> str:
        cols, rows = await _build_shipment_rows_for_user(user_id, ids)
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(cols)
        for r in rows:
            writer.writerow(r)
        # Prepend a UTF-8 BOM so Numbers / LibreOffice / Excel 2016+
        # open the download with the correct encoding. Pre-2016
        # Excel on Windows falls back to Windows-1252 for CSV — for
        # those users, direct them to the XLSX endpoint instead.
        return "\ufeff" + buf.getvalue()

    # ─── XLSX builder ───────────────────────────────────────────────
    #
    # Root cause for the "garbage Gujarati/Hindi names in Excel" bug:
    # Windows Excel < 2016 ignores the UTF-8 BOM in .csv files and
    # reads them using the system's ANSI code page (Windows-1252 on
    # IN/UK/US locales). UTF-8 encoded Indic characters then render
    # as sequences like "àn àn¥‹àn'".
    #
    # .xlsx is a ZIP of UTF-8 XML — the encoding is *baked into the
    # file format itself* — so Excel of any version renders every
    # script correctly. We keep the CSV path intact (Numbers /
    # LibreOffice / modern Excel handle it fine) and expose an
    # "Excel (.xlsx)" button alongside the existing CSV button.
    def _build_xlsx_bytes(
        columns: List[str], rows: List[List[Any]],
    ) -> bytes:
        # Lazy import — openpyxl loads a fair amount of XML machinery.
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Shipments"

        # Header row with a subtle bold formatting.
        header_font = Font(bold=True)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Data rows — coerce everything to str-safe values; openpyxl
        # supports Indic strings out of the box.
        for r in rows:
            ws.append([("" if v is None else v) for v in r])

        # Freeze the header + auto-size the first 10 columns to a
        # readable width; skipping the full auto-fit pass to keep
        # generation fast on large exports.
        ws.freeze_panes = "A2"
        for i, col_name in enumerate(columns[:10], start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(
                12, min(30, len(str(col_name)) + 4)
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    _XLSX_MEDIA = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    _XLSX_HEADERS = {
        "Content-Disposition": 'attachment; filename="shipments_export.xlsx"',
    }

    # ═══════════════════════════════════════════════════════════════
    #                          ROUTES
    # ═══════════════════════════════════════════════════════════════

    @shipments_export_router.get(
        "/shipments/export/csv", response_class=PlainTextResponse,
    )
    async def export_csv(
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        # Legacy GET → no filter, export everything for the user.
        body = await _build_csv_for_user(current_user["id"], None)
        return PlainTextResponse(body, media_type="text/csv")

    @shipments_export_router.post(
        "/shipments/export/csv", response_class=PlainTextResponse,
    )
    async def export_csv_filtered(
        payload: _ExportCsvBody,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        # POST lets the frontend send the exact id list visible on
        # screen (after applying status + date filters).
        body = await _build_csv_for_user(current_user["id"], payload.ids)
        return PlainTextResponse(body, media_type="text/csv")

    @shipments_export_router.get("/shipments/export/xlsx")
    async def export_xlsx(
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        cols, rows = await _build_shipment_rows_for_user(
            current_user["id"], None,
        )
        data = _build_xlsx_bytes(cols, rows)
        return Response(
            content=data,
            media_type=_XLSX_MEDIA,
            headers=_XLSX_HEADERS,
        )

    @shipments_export_router.post("/shipments/export/xlsx")
    async def export_xlsx_filtered(
        payload: _ExportCsvBody,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        cols, rows = await _build_shipment_rows_for_user(
            current_user["id"], payload.ids,
        )
        data = _build_xlsx_bytes(cols, rows)
        return Response(
            content=data,
            media_type=_XLSX_MEDIA,
            headers=_XLSX_HEADERS,
        )

    _logger.info(
        "shipments_export router mounted: 4 endpoints (csv GET/POST + xlsx GET/POST)",
    )
