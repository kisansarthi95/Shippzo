"""
Courier Billing Report — Phase 2.5
══════════════════════════════════════════════════════════════════
Aggregates each user's own shipments by courier_partner over a
chosen period and emits two surfaces:

  1. JSON for the in-app Reports screen (KPI cards + detail table).
  2. Multi-sheet .xlsx for one-tap download / share with the courier
     partner directly from WhatsApp / email.

Data source: the `shipments` collection. Phase 2B introduced the per-
shipment fields (rate_applied, rate_basis, package_type, category)
that this report relies on. Older rows have rate_applied=0 — they're
counted in volume but contribute ₹0 to charges; the UI surfaces a
`rows_without_rate` KPI so the user can fix them up manually.
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse


reports_router = APIRouter(prefix="/api", tags=["reports"])


# ──────────────────────────────────────────────────────────────────
# Helpers (pure-fn — no DB / auth dependency)
# ──────────────────────────────────────────────────────────────────


def _resolve_period(
    range_key: str,
    from_iso: Optional[str],
    to_iso: Optional[str],
) -> Dict[str, Any]:
    """Map UI range chips to (from, to) datetimes in UTC. The UI may
    also send `from` / `to` for full custom ranges — those win when
    range_key=='custom'."""
    now = datetime.now(timezone.utc)
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)

    def fmt_label(start: datetime, end: datetime) -> str:
        if start.year == end.year and start.month == end.month:
            return start.strftime("%B %Y")
        return f"{start.strftime('%-d %b')} – {end.strftime('%-d %b %Y')}"

    if range_key == "custom" and from_iso and to_iso:
        try:
            start = datetime.fromisoformat(from_iso.replace("Z", "+00:00"))
            end   = datetime.fromisoformat(to_iso.replace("Z", "+00:00"))
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(400, f"Invalid custom range: {exc}") from exc
    elif range_key == "this_week":
        start = today - timedelta(days=today.weekday())
        end   = today + timedelta(days=1)
    elif range_key == "last_week":
        end   = today - timedelta(days=today.weekday())
        start = end - timedelta(days=7)
    elif range_key == "last_month":
        first_this = today.replace(day=1)
        end        = first_this
        start      = (first_this - timedelta(days=1)).replace(day=1)
    elif range_key == "last_30":
        start = today - timedelta(days=30)
        end   = today + timedelta(days=1)
    else:  # default: this_month
        start = today.replace(day=1)
        end   = today + timedelta(days=1)

    return {
        "from":  start.isoformat(),
        "to":    end.isoformat(),
        "label": fmt_label(start, end - timedelta(seconds=1)),
    }


def _bucket_payment(pm: Optional[str]) -> str:
    norm = (pm or "").strip().upper()
    if norm == "COD":
        return "COD"
    if norm in ("PREPAID", "PAID"):
        return "PREPAID"
    return "Other"


# Phase 2.5 — Weight bucket ordering used by both the aggregator and
# the Excel renderer; defined at module scope so the latter (which
# lives outside the `init()` closure) can see it.
_WEIGHT_ORDER = ["0–500 g", "500 g–1 kg", "1–2 kg", "2–5 kg", "5 kg+", "Unknown"]


# ──────────────────────────────────────────────────────────────────
# init() — server.py calls this on startup so we can close over the
# late-bound `db` handle and the `get_current_user` dependency.
# ──────────────────────────────────────────────────────────────────


def init() -> None:
    """Late-bind helpers from server.py and register the routes inside
    the closure so FastAPI sees real dependencies (not None)."""
    from server import db, get_current_user  # noqa: WPS433

    async def _resolve_token_user(token: str) -> Optional[Dict[str, Any]]:
        """Decode the bearer token via auth.py and return the user dict.
        Used as a fallback for the Excel route which the browser hits
        without an Authorization header."""
        try:
            from auth import decode_token as _decode_token  # type: ignore[attr-defined]
            payload = _decode_token(token)
            uid = payload.get("sub") or payload.get("user_id") or payload.get("id")
            if not uid:
                return None
            return await db.users.find_one({"id": uid}, {"_id": 0})
        except Exception:
            return None

    async def _aggregate(
        user_id: str,
        range_key: str,
        courier_id: Optional[str],
        from_iso: Optional[str],
        to_iso: Optional[str],
    ) -> Dict[str, Any]:
        period = _resolve_period(range_key, from_iso, to_iso)
        match: Dict[str, Any] = {
            "user_id":    user_id,
            "deleted_at": {"$exists": False},
            "created_at": {"$gte": period["from"], "$lt": period["to"]},
        }
        if courier_id and courier_id != "all":
            match["courier_id"] = courier_id

        cursor = db.shipments.find(match, {
            "_id": 0, "id": 1, "courier_id": 1, "courier_name": 1,
            "tracking_id": 1, "order_id": 1, "customer_name": 1,
            "customer_phone": 1, "city": 1, "state": 1, "weight": 1,
            "amount": 1, "payment_mode": 1, "status": 1, "created_at": 1,
            "rate_applied": 1, "rate_basis": 1, "package_type": 1,
            "category": 1, "variant_name": 1,
        }).sort("created_at", 1)
        rows: List[Dict[str, Any]] = await cursor.to_list(5000)

        by_courier: Dict[str, Dict[str, Any]] = {}
        rows_without_rate = 0
        grand_charges = 0.0
        grand_count = 0

        for r in rows:
            cid = r.get("courier_id") or ""
            cname = r.get("courier_name") or "Unknown"
            bucket_id = cid or cname
            bucket = by_courier.setdefault(bucket_id, {
                "courier_id":      cid,
                "courier_name":    cname,
                "total_shipments": 0,
                "total_charges":   0.0,
                "cod":            {"count": 0, "amount": 0.0},
                "prepaid":        {"count": 0, "amount": 0.0},
                "other":          {"count": 0, "amount": 0.0},
                "by_package_type": {},
                "by_state":        {},
                "shipments":       [],
            })

            rate = float(r.get("rate_applied") or 0)
            if rate <= 0:
                rows_without_rate += 1
            bucket["total_shipments"] += 1
            bucket["total_charges"]   += rate
            grand_count   += 1
            grand_charges += rate

            bucket_key = _bucket_payment(r.get("payment_mode"))
            target = bucket["cod"] if bucket_key == "COD" else (
                     bucket["prepaid"] if bucket_key == "PREPAID" else bucket["other"])
            target["count"]  += 1
            target["amount"] += rate

            pt = (r.get("package_type") or "—").strip() or "—"
            pt_b = bucket["by_package_type"].setdefault(pt, {"count": 0, "amount": 0.0})
            pt_b["count"]  += 1
            pt_b["amount"] += rate

            st = (r.get("state") or "—").strip() or "—"
            st_b = bucket["by_state"].setdefault(st, {"count": 0})
            st_b["count"] += 1

            bucket["shipments"].append({
                "id":            r.get("id"),
                "tracking_id":   r.get("tracking_id"),
                "order_id":      r.get("order_id"),
                "date":          (r.get("created_at") or "")[:10],
                "customer_name": r.get("customer_name"),
                "city":          r.get("city"),
                "state":         r.get("state"),
                "weight":        r.get("weight"),
                "rate":          rate,
                "payment_mode":  bucket_key,
                "status":        r.get("status"),
                "package_type":  r.get("package_type"),
                "variant_name":  r.get("variant_name"),
            })

        couriers_out: List[Dict[str, Any]] = []
        for b in by_courier.values():
            b["by_package_type"] = sorted(
                ({"type": k, **v} for k, v in b["by_package_type"].items()),
                key=lambda x: x["count"], reverse=True,
            )
            b["by_state"] = sorted(
                ({"state": k, **v} for k, v in b["by_state"].items()),
                key=lambda x: x["count"], reverse=True,
            )
            couriers_out.append(b)
        couriers_out.sort(key=lambda x: x["total_charges"], reverse=True)

        return {
            "period":   period,
            "couriers": couriers_out,
            "grand_total": {
                "shipments": grand_count,
                "charges":   grand_charges,
            },
            "rows_without_rate": rows_without_rate,
        }

    @reports_router.get("/me/reports/courier-billing")
    async def courier_billing_report(
        range_key: str = Query("this_month", alias="range"),
        courier_id: Optional[str] = Query(None),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        current_user: Dict[str, Any] = Depends(get_current_user),
    ):
        return await _aggregate(
            current_user["id"], range_key, courier_id, from_iso, to_iso,
        )

    @reports_router.get("/me/reports/courier-billing/excel")
    async def courier_billing_report_excel(
        request: "Request",  # noqa: F821 — string-typed for forward import
        range_key: str = Query("this_month", alias="range"),
        courier_id: Optional[str] = Query(None),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        token: Optional[str] = Query(None),
    ):
        # Prefer the standard Authorization header, fall back to the
        # ?token= query carrier so that browser-initiated downloads
        # (which can't set custom headers) still work.
        current_user: Optional[Dict[str, Any]] = None
        bearer = request.headers.get("authorization") or ""
        if bearer.lower().startswith("bearer "):
            current_user = await _resolve_token_user(bearer.split(" ", 1)[1])
        if current_user is None and token:
            current_user = await _resolve_token_user(token)
        if current_user is None:
            raise HTTPException(401, "Auth required")
        payload = await _aggregate(
            current_user["id"], range_key, courier_id, from_iso, to_iso,
        )
        return _build_excel(payload, current_user)

    # ──────────────────────────────────────────────────────────────
    # Phase 2.5 — 3 additional reports share the same date-range
    # selector + Excel download pattern. They each pull from the
    # `shipments` collection but slice the data along a different
    # dimension:
    #   • returns        — focuses on status="Returned" rows.
    #   • weight-wise    — buckets by parsed weight (g/kg).
    #   • partner-comp   — side-by-side metrics across all couriers.
    #   • reconciliation — expected vs received COD per courier.
    # ──────────────────────────────────────────────────────────────

    async def _fetch_shipments(
        user_id: str, range_key: str,
        from_iso: Optional[str], to_iso: Optional[str],
    ):
        period = _resolve_period(range_key, from_iso, to_iso)
        match = {
            "user_id":    user_id,
            "deleted_at": {"$exists": False},
            "created_at": {"$gte": period["from"], "$lt": period["to"]},
        }
        rows = await db.shipments.find(match, {
            "_id": 0, "id": 1, "courier_id": 1, "courier_name": 1,
            "tracking_id": 1, "order_id": 1, "customer_name": 1,
            "customer_phone": 1, "city": 1, "state": 1, "weight": 1,
            "amount": 1, "cod_amount": 1, "payment_mode": 1,
            "status": 1, "created_at": 1, "delivered_at": 1,
            "rate_applied": 1, "package_type": 1, "category": 1,
            "return_reason": 1, "cod_received": 1, "payment_received": 1,
        }).sort("created_at", 1).to_list(8000)
        return period, rows

    def _parse_weight_grams(raw) -> Optional[float]:
        """Best-effort weight → grams. Accepts '500g', '1.5 kg', '2', etc."""
        if raw is None:
            return None
        s = str(raw).strip().lower().replace(" ", "")
        if not s:
            return None
        try:
            if s.endswith("kg"):
                return float(s[:-2]) * 1000.0
            if s.endswith("g"):
                return float(s[:-1])
            # bare number → assume grams under 50, kg above
            n = float(s)
            return n * 1000.0 if n < 50 else n
        except ValueError:
            return None

    def _weight_bucket(g: Optional[float]) -> str:
        if g is None: return "Unknown"
        if g <= 500:  return "0–500 g"
        if g <= 1000: return "500 g–1 kg"
        if g <= 2000: return "1–2 kg"
        if g <= 5000: return "2–5 kg"
        return "5 kg+"

    _WEIGHT_ORDER = ["0–500 g", "500 g–1 kg", "1–2 kg", "2–5 kg", "5 kg+", "Unknown"]

    # ─── Return Analysis ─────────────────────────────────────────
    async def _returns_aggregate(user_id, range_key, from_iso, to_iso):
        period, rows = await _fetch_shipments(user_id, range_key, from_iso, to_iso)
        total = len(rows)
        returns = [r for r in rows if (r.get("status") or "").lower() == "returned"]
        return_rate = (len(returns) / total * 100.0) if total else 0.0

        by_courier: Dict[str, Dict[str, Any]] = {}
        by_reason:  Dict[str, int] = {}
        by_customer: Dict[str, Dict[str, Any]] = {}
        rows_out: List[Dict[str, Any]] = []
        for r in rows:
            cn = r.get("courier_name") or "Unknown"
            b = by_courier.setdefault(cn, {"courier_name": cn, "total": 0, "returned": 0})
            b["total"] += 1
        for r in returns:
            cn = r.get("courier_name") or "Unknown"
            b = by_courier.setdefault(cn, {"courier_name": cn, "total": 0, "returned": 0})
            b["returned"] += 1
            reason = (r.get("return_reason") or "Not specified").strip() or "Not specified"
            by_reason[reason] = by_reason.get(reason, 0) + 1
            cust = r.get("customer_phone") or r.get("customer_name") or "Unknown"
            cb = by_customer.setdefault(cust, {
                "name": r.get("customer_name") or "—",
                "phone": r.get("customer_phone") or "",
                "count": 0,
            })
            cb["count"] += 1
            rows_out.append({
                "id": r.get("id"),
                "tracking_id": r.get("tracking_id"),
                "order_id": r.get("order_id"),
                "customer": r.get("customer_name"),
                "phone": r.get("customer_phone"),
                "courier": r.get("courier_name"),
                "city": r.get("city"),
                "state": r.get("state"),
                "amount": float(r.get("amount") or 0),
                "payment_mode": (r.get("payment_mode") or "—").upper(),
                "reason": reason,
                "date": (r.get("created_at") or "")[:10],
            })
        couriers_out = []
        for b in by_courier.values():
            b["return_rate"] = round((b["returned"] / b["total"] * 100.0), 2) if b["total"] else 0.0
            couriers_out.append(b)
        couriers_out.sort(key=lambda x: x["returned"], reverse=True)
        reasons_out = sorted(
            [{"reason": k, "count": v} for k, v in by_reason.items()],
            key=lambda x: x["count"], reverse=True,
        )
        repeat_customers = sorted(
            [c for c in by_customer.values() if c["count"] >= 2],
            key=lambda x: x["count"], reverse=True,
        )[:25]
        return {
            "period":   period,
            "summary":  {
                "total_shipments":   total,
                "total_returns":     len(returns),
                "return_rate":       round(return_rate, 2),
                "unique_customers":  len({(r.get("customer_phone") or r.get("customer_name") or "") for r in returns}),
            },
            "by_courier":      couriers_out,
            "by_reason":       reasons_out,
            "repeat_customers": repeat_customers,
            "returns":          rows_out,
        }

    @reports_router.get("/me/reports/return-analysis")
    async def return_analysis_report(
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        current_user: Dict[str, Any] = Depends(get_current_user),
    ):
        return await _returns_aggregate(current_user["id"], range_key, from_iso, to_iso)

    @reports_router.get("/me/reports/return-analysis/excel")
    async def return_analysis_excel(
        request: "Request",  # noqa: F821
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        token: Optional[str] = Query(None),
    ):
        current_user = None
        bearer = request.headers.get("authorization") or ""
        if bearer.lower().startswith("bearer "):
            current_user = await _resolve_token_user(bearer.split(" ", 1)[1])
        if current_user is None and token:
            current_user = await _resolve_token_user(token)
        if current_user is None:
            raise HTTPException(401, "Auth required")
        payload = await _returns_aggregate(current_user["id"], range_key, from_iso, to_iso)
        return _build_returns_excel(payload, current_user)

    # ─── Weight-wise Breakup ─────────────────────────────────────
    async def _weight_aggregate(user_id, range_key, from_iso, to_iso):
        period, rows = await _fetch_shipments(user_id, range_key, from_iso, to_iso)
        buckets: Dict[str, Dict[str, Any]] = {b: {
            "bucket": b, "count": 0, "revenue": 0.0, "rate_total": 0.0,
        } for b in _WEIGHT_ORDER}
        by_courier_weight: Dict[str, Dict[str, int]] = {}
        for r in rows:
            g = _parse_weight_grams(r.get("weight"))
            bk = _weight_bucket(g)
            b = buckets[bk]
            b["count"] += 1
            b["revenue"] += float(r.get("amount") or 0)
            b["rate_total"] += float(r.get("rate_applied") or 0)
            cn = r.get("courier_name") or "Unknown"
            by_courier_weight.setdefault(cn, {b: 0 for b in _WEIGHT_ORDER})
            by_courier_weight[cn][bk] = by_courier_weight[cn].get(bk, 0) + 1
        # avg rate per bucket
        out_buckets = []
        for b in _WEIGHT_ORDER:
            row = buckets[b]
            row["avg_cost"] = round(row["rate_total"] / row["count"], 2) if row["count"] else 0
            out_buckets.append(row)
        couriers_out = [
            {"courier_name": k, "by_bucket": v, "total": sum(v.values())}
            for k, v in by_courier_weight.items()
        ]
        couriers_out.sort(key=lambda x: x["total"], reverse=True)
        return {
            "period":   period,
            "buckets":  out_buckets,
            "couriers": couriers_out,
            "total_shipments": len(rows),
        }

    @reports_router.get("/me/reports/weight-wise")
    async def weight_wise_report(
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        current_user: Dict[str, Any] = Depends(get_current_user),
    ):
        return await _weight_aggregate(current_user["id"], range_key, from_iso, to_iso)

    @reports_router.get("/me/reports/weight-wise/excel")
    async def weight_wise_excel(
        request: "Request",  # noqa: F821
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        token: Optional[str] = Query(None),
    ):
        current_user = None
        bearer = request.headers.get("authorization") or ""
        if bearer.lower().startswith("bearer "):
            current_user = await _resolve_token_user(bearer.split(" ", 1)[1])
        if current_user is None and token:
            current_user = await _resolve_token_user(token)
        if current_user is None:
            raise HTTPException(401, "Auth required")
        payload = await _weight_aggregate(current_user["id"], range_key, from_iso, to_iso)
        return _build_weight_excel(payload, current_user)

    # ─── Partner Comparison ──────────────────────────────────────
    async def _partner_aggregate(user_id, range_key, from_iso, to_iso):
        period, rows = await _fetch_shipments(user_id, range_key, from_iso, to_iso)
        by: Dict[str, Dict[str, Any]] = {}
        for r in rows:
            cn = r.get("courier_name") or "Unknown"
            b = by.setdefault(cn, {
                "courier_name": cn, "total": 0, "delivered": 0, "returned": 0,
                "pending": 0, "revenue": 0.0, "cost": 0.0, "cod": 0.0,
            })
            b["total"] += 1
            st = (r.get("status") or "").lower()
            if st == "delivered":   b["delivered"] += 1
            elif st == "returned":  b["returned"] += 1
            else:                   b["pending"]   += 1
            b["revenue"] += float(r.get("amount") or 0)
            b["cost"]    += float(r.get("rate_applied") or 0)
            if (r.get("payment_mode") or "").upper() == "COD":
                b["cod"] += float(r.get("cod_amount") or r.get("amount") or 0)
        out = []
        for b in by.values():
            tot = b["total"] or 1
            b["delivery_rate"] = round(b["delivered"] / tot * 100.0, 2)
            b["return_rate"]   = round(b["returned"]  / tot * 100.0, 2)
            b["avg_cost"]      = round(b["cost"] / tot, 2)
            b["margin"]        = round(b["revenue"] - b["cost"], 2)
            out.append(b)
        out.sort(key=lambda x: x["total"], reverse=True)
        # best/worst rankings (only when there's data)
        ranks = {"best_delivery": None, "worst_delivery": None,
                 "best_returns":  None, "worst_returns":  None,
                 "cheapest":       None, "highest_revenue": None}
        if out:
            ranks["best_delivery"]   = max(out, key=lambda x: x["delivery_rate"])["courier_name"]
            ranks["worst_delivery"]  = min(out, key=lambda x: x["delivery_rate"])["courier_name"]
            ranks["best_returns"]    = min(out, key=lambda x: x["return_rate"])["courier_name"]
            ranks["worst_returns"]   = max(out, key=lambda x: x["return_rate"])["courier_name"]
            ranks["cheapest"]        = min(out, key=lambda x: x["avg_cost"] if x["avg_cost"] else 1e9)["courier_name"]
            ranks["highest_revenue"] = max(out, key=lambda x: x["revenue"])["courier_name"]
        return {
            "period":   period,
            "couriers": out,
            "rankings": ranks,
            "total_shipments": sum(c["total"] for c in out),
        }

    @reports_router.get("/me/reports/partner-comparison")
    async def partner_comparison_report(
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        current_user: Dict[str, Any] = Depends(get_current_user),
    ):
        return await _partner_aggregate(current_user["id"], range_key, from_iso, to_iso)

    @reports_router.get("/me/reports/partner-comparison/excel")
    async def partner_comparison_excel(
        request: "Request",  # noqa: F821
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        token: Optional[str] = Query(None),
    ):
        current_user = None
        bearer = request.headers.get("authorization") or ""
        if bearer.lower().startswith("bearer "):
            current_user = await _resolve_token_user(bearer.split(" ", 1)[1])
        if current_user is None and token:
            current_user = await _resolve_token_user(token)
        if current_user is None:
            raise HTTPException(401, "Auth required")
        payload = await _partner_aggregate(current_user["id"], range_key, from_iso, to_iso)
        return _build_partner_excel(payload, current_user)

    # ─── Reconciliation (COD) ───────────────────────────────────
    async def _recon_aggregate(user_id, range_key, from_iso, to_iso):
        period, rows = await _fetch_shipments(user_id, range_key, from_iso, to_iso)
        by: Dict[str, Dict[str, Any]] = {}
        delivered_cod_rows = []
        for r in rows:
            if (r.get("payment_mode") or "").upper() != "COD":
                continue
            st = (r.get("status") or "").lower()
            cn = r.get("courier_name") or "Unknown"
            amt = float(r.get("cod_amount") or r.get("amount") or 0)
            b = by.setdefault(cn, {
                "courier_name":     cn,
                "delivered_count":  0, "delivered_amt":   0.0,
                "received_count":   0, "received_amt":    0.0,
                "pending_count":    0, "pending_amt":     0.0,
            })
            received = bool(r.get("cod_received") or r.get("payment_received"))
            if st == "delivered":
                b["delivered_count"] += 1
                b["delivered_amt"]   += amt
                if received:
                    b["received_count"] += 1
                    b["received_amt"]   += amt
                else:
                    b["pending_count"]  += 1
                    b["pending_amt"]    += amt
                    delivered_cod_rows.append({
                        "id": r.get("id"),
                        "tracking_id": r.get("tracking_id"),
                        "order_id": r.get("order_id"),
                        "customer": r.get("customer_name"),
                        "courier": cn,
                        "amount": amt,
                        "delivered_at": (r.get("delivered_at") or r.get("created_at") or "")[:10],
                    })
        out = sorted(by.values(), key=lambda x: x["pending_amt"], reverse=True)
        totals = {
            "delivered_count": sum(b["delivered_count"] for b in out),
            "delivered_amt":   round(sum(b["delivered_amt"]   for b in out), 2),
            "received_count":  sum(b["received_count"]  for b in out),
            "received_amt":    round(sum(b["received_amt"]    for b in out), 2),
            "pending_count":   sum(b["pending_count"]   for b in out),
            "pending_amt":     round(sum(b["pending_amt"]     for b in out), 2),
        }
        return {
            "period":   period,
            "couriers": out,
            "totals":   totals,
            "pending":  delivered_cod_rows[:200],
        }

    @reports_router.get("/me/reports/reconciliation")
    async def reconciliation_report(
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        current_user: Dict[str, Any] = Depends(get_current_user),
    ):
        return await _recon_aggregate(current_user["id"], range_key, from_iso, to_iso)

    @reports_router.get("/me/reports/reconciliation/excel")
    async def reconciliation_excel(
        request: "Request",  # noqa: F821
        range_key: str = Query("this_month", alias="range"),
        from_iso: Optional[str] = Query(None, alias="from"),
        to_iso: Optional[str] = Query(None, alias="to"),
        token: Optional[str] = Query(None),
    ):
        current_user = None
        bearer = request.headers.get("authorization") or ""
        if bearer.lower().startswith("bearer "):
            current_user = await _resolve_token_user(bearer.split(" ", 1)[1])
        if current_user is None and token:
            current_user = await _resolve_token_user(token)
        if current_user is None:
            raise HTTPException(401, "Auth required")
        payload = await _recon_aggregate(current_user["id"], range_key, from_iso, to_iso)
        return _build_recon_excel(payload, current_user)


def _build_excel(payload: Dict[str, Any], current_user: Dict[str, Any]) -> StreamingResponse:
    """Render the multi-sheet workbook from an aggregated payload.
    Sheet 1 is a roll-up summary; sheets 2+ are per-courier bills."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4FBF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(bold=True, size=14, color="1F4FBF")
    LABEL_FONT  = Font(bold=True, size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")

    business_name = (
        current_user.get("shop_name") or current_user.get("name") or "—"
    )

    # Summary sheet --------------------------------------------------
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = f"Courier Billing — {payload['period']['label']}"
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A1:E1")
    summary["A2"] = f"Business: {business_name}"
    summary["A2"].font = LABEL_FONT
    summary.merge_cells("A2:E2")
    summary["A3"] = (
        f"Total shipments: {payload['grand_total']['shipments']}    "
        f"Total charges: ₹{payload['grand_total']['charges']:.0f}"
    )
    summary["A3"].font = LABEL_FONT
    summary.merge_cells("A3:E3")
    if payload["rows_without_rate"]:
        summary["A4"] = (
            f"⚠️ {payload['rows_without_rate']} shipment(s) have no rate set "
            "— they count in volume but contribute ₹0 to charges."
        )
        summary.merge_cells("A4:E4")
        start_row = 6
    else:
        start_row = 5

    headers = ["Courier", "Shipments", "COD ₹", "Prepaid ₹", "Total ₹"]
    for col, h in enumerate(headers, 1):
        cell = summary.cell(row=start_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for i, c in enumerate(payload["couriers"], 1):
        r = start_row + i
        summary.cell(row=r, column=1, value=c["courier_name"])
        summary.cell(row=r, column=2, value=c["total_shipments"])
        summary.cell(row=r, column=3, value=round(c["cod"]["amount"]))
        summary.cell(row=r, column=4, value=round(c["prepaid"]["amount"]))
        tot = summary.cell(row=r, column=5, value=round(c["total_charges"]))
        tot.font = Font(bold=True)
    g = start_row + len(payload["couriers"]) + 1
    summary.cell(row=g, column=1, value="GRAND TOTAL").font = Font(bold=True)
    summary.cell(row=g, column=2, value=payload["grand_total"]["shipments"]).font = Font(bold=True)
    summary.cell(row=g, column=5, value=round(payload["grand_total"]["charges"])).font = Font(bold=True)
    for col in range(1, 6):
        summary.column_dimensions[get_column_letter(col)].width = 18

    # Per-courier sheets --------------------------------------------
    for c in payload["couriers"]:
        safe = "".join(ch for ch in (c["courier_name"] or "Courier") if ch not in r'[]:*?/\\')[:28]
        ws = wb.create_sheet(title=safe or "Courier")
        ws["A1"] = f"Courier Bill — {c['courier_name']}"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:H1")
        ws["A2"] = (
            f"Period: {payload['period']['label']}    "
            f"Shipments: {c['total_shipments']}    "
            f"Total: ₹{c['total_charges']:.0f}"
        )
        ws["A2"].font = LABEL_FONT
        ws.merge_cells("A2:H2")
        ws["A3"] = (
            f"COD: {c['cod']['count']} (₹{c['cod']['amount']:.0f})   "
            f"Prepaid: {c['prepaid']['count']} (₹{c['prepaid']['amount']:.0f})"
        )
        ws.merge_cells("A3:H3")

        ship_headers = [
            "Date", "Tracking ID", "Order ID", "Customer", "City", "State",
            "Weight", "Pkg Type", "Pay Mode", "Rate ₹",
        ]
        for col, h in enumerate(ship_headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        for i, s in enumerate(c["shipments"], 1):
            r = 5 + i
            ws.cell(row=r, column=1,  value=s.get("date"))
            ws.cell(row=r, column=2,  value=s.get("tracking_id") or "")
            ws.cell(row=r, column=3,  value=s.get("order_id") or "")
            ws.cell(row=r, column=4,  value=s.get("customer_name") or "")
            ws.cell(row=r, column=5,  value=s.get("city") or "")
            ws.cell(row=r, column=6,  value=s.get("state") or "")
            ws.cell(row=r, column=7,  value=s.get("weight") or "")
            ws.cell(row=r, column=8,  value=s.get("package_type") or "")
            ws.cell(row=r, column=9,  value=s.get("payment_mode") or "")
            ws.cell(row=r, column=10, value=round(s.get("rate") or 0))

        last = 5 + len(c["shipments"]) + 1
        ws.cell(row=last, column=9,  value="Total").font = Font(bold=True)
        ws.cell(row=last, column=10, value=round(c["total_charges"])).font = Font(bold=True)

        widths = [11, 18, 12, 22, 14, 14, 10, 14, 11, 11]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Latin-1 only allowed in Content-Disposition — strip Unicode dashes
    # and any other non-ASCII so the period label is safe in the header.
    raw = payload["period"]["label"].replace("–", "-").replace("—", "-")
    fname_period = "".join(ch if ord(ch) < 128 else "_" for ch in raw).replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="CourierBill_{fname_period}.xlsx"',
        },
    )



# ─── Phase 2.5 — Excel builders for the 4 additional reports ───────────
# All four share the same chrome (header colours, period label, business
# name, ascii-safe filename) so the workbooks feel like a coherent suite.

def _xlsx_chrome():
    """One-shot import + style-set used by every report workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    return {
        "Workbook":         Workbook,
        "Alignment":        Alignment,
        "Font":             Font,
        "PatternFill":      PatternFill,
        "get_column_letter": get_column_letter,
        "HEADER_FILL":      PatternFill("solid", fgColor="1F4FBF"),
        "HEADER_FONT":      Font(bold=True, color="FFFFFF", size=11),
        "TITLE_FONT":       Font(bold=True, size=14, color="1F4FBF"),
        "LABEL_FONT":       Font(bold=True, size=10),
        "CENTER":           Alignment(horizontal="center", vertical="center"),
    }


def _xlsx_response(buf, *, prefix: str, label: str) -> StreamingResponse:
    raw = (label or "").replace("–", "-").replace("—", "-")
    fname = "".join(ch if ord(ch) < 128 else "_" for ch in raw).replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{prefix}_{fname}.xlsx"',
        },
    )


def _build_returns_excel(payload, current_user) -> StreamingResponse:
    c = _xlsx_chrome()
    wb = c["Workbook"]()
    biz = current_user.get("shop_name") or current_user.get("name") or "—"

    s = wb.active
    s.title = "Summary"
    s["A1"] = f"Return Analysis — {payload['period']['label']}"
    s["A1"].font = c["TITLE_FONT"]; s.merge_cells("A1:E1")
    s["A2"] = f"Business: {biz}"; s["A2"].font = c["LABEL_FONT"]; s.merge_cells("A2:E2")
    sm = payload["summary"]
    s["A3"] = (f"Total: {sm['total_shipments']}    Returns: {sm['total_returns']}    "
               f"Return rate: {sm['return_rate']}%    Unique customers: {sm['unique_customers']}")
    s["A3"].font = c["LABEL_FONT"]; s.merge_cells("A3:E3")

    # By courier
    s["A5"] = "By Courier"; s["A5"].font = c["TITLE_FONT"]
    for col, h in enumerate(["Courier", "Total", "Returned", "Return %"], 1):
        cell = s.cell(row=6, column=col, value=h)
        cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
    for i, r in enumerate(payload["by_courier"], 1):
        s.cell(row=6+i, column=1, value=r["courier_name"])
        s.cell(row=6+i, column=2, value=r["total"])
        s.cell(row=6+i, column=3, value=r["returned"])
        s.cell(row=6+i, column=4, value=f"{r['return_rate']}%")

    # By reason
    base = 6 + len(payload["by_courier"]) + 3
    s.cell(row=base-1, column=1, value="By Reason").font = c["TITLE_FONT"]
    for col, h in enumerate(["Reason", "Count"], 1):
        cell = s.cell(row=base, column=col, value=h)
        cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
    for i, r in enumerate(payload["by_reason"], 1):
        s.cell(row=base+i, column=1, value=r["reason"])
        s.cell(row=base+i, column=2, value=r["count"])
    for col in range(1, 6):
        s.column_dimensions[c["get_column_letter"](col)].width = 22

    # Returned shipments detail sheet
    if payload["returns"]:
        ws = wb.create_sheet(title="Returned Shipments")
        for col, h in enumerate(
            ["Date", "Tracking ID", "Order ID", "Customer", "Phone",
             "Courier", "City", "State", "Pay Mode", "Amount", "Reason"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
        for i, r in enumerate(payload["returns"], 1):
            row = i + 1
            ws.cell(row=row, column=1, value=r.get("date"))
            ws.cell(row=row, column=2, value=r.get("tracking_id"))
            ws.cell(row=row, column=3, value=r.get("order_id"))
            ws.cell(row=row, column=4, value=r.get("customer"))
            ws.cell(row=row, column=5, value=r.get("phone"))
            ws.cell(row=row, column=6, value=r.get("courier"))
            ws.cell(row=row, column=7, value=r.get("city"))
            ws.cell(row=row, column=8, value=r.get("state"))
            ws.cell(row=row, column=9, value=r.get("payment_mode"))
            ws.cell(row=row, column=10, value=round(r.get("amount") or 0))
            ws.cell(row=row, column=11, value=r.get("reason"))
        for col in range(1, 12):
            ws.column_dimensions[c["get_column_letter"](col)].width = 15

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return _xlsx_response(buf, prefix="ReturnAnalysis", label=payload["period"]["label"])


def _build_weight_excel(payload, current_user) -> StreamingResponse:
    c = _xlsx_chrome()
    wb = c["Workbook"]()
    biz = current_user.get("shop_name") or current_user.get("name") or "—"

    s = wb.active
    s.title = "Weight Buckets"
    s["A1"] = f"Weight-wise Breakup — {payload['period']['label']}"
    s["A1"].font = c["TITLE_FONT"]; s.merge_cells("A1:E1")
    s["A2"] = f"Business: {biz}    Total shipments: {payload['total_shipments']}"
    s["A2"].font = c["LABEL_FONT"]; s.merge_cells("A2:E2")

    for col, h in enumerate(["Weight Bucket", "Count", "Revenue ₹", "Avg Cost ₹"], 1):
        cell = s.cell(row=4, column=col, value=h)
        cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
    for i, b in enumerate(payload["buckets"], 1):
        s.cell(row=4+i, column=1, value=b["bucket"])
        s.cell(row=4+i, column=2, value=b["count"])
        s.cell(row=4+i, column=3, value=round(b["revenue"]))
        s.cell(row=4+i, column=4, value=round(b["avg_cost"]))
    for col in range(1, 5):
        s.column_dimensions[c["get_column_letter"](col)].width = 18

    # Per-courier sheet
    ws = wb.create_sheet(title="By Courier")
    cols_header = ["Courier"] + _WEIGHT_ORDER + ["Total"]
    for col, h in enumerate(cols_header, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
    for i, row in enumerate(payload["couriers"], 1):
        ws.cell(row=1+i, column=1, value=row["courier_name"])
        for j, bk in enumerate(_WEIGHT_ORDER, 2):
            ws.cell(row=1+i, column=j, value=row["by_bucket"].get(bk, 0))
        ws.cell(row=1+i, column=len(cols_header), value=row["total"])
    for col in range(1, len(cols_header) + 1):
        ws.column_dimensions[c["get_column_letter"](col)].width = 14

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return _xlsx_response(buf, prefix="WeightBreakup", label=payload["period"]["label"])


def _build_partner_excel(payload, current_user) -> StreamingResponse:
    c = _xlsx_chrome()
    wb = c["Workbook"]()
    biz = current_user.get("shop_name") or current_user.get("name") or "—"

    s = wb.active
    s.title = "Comparison"
    s["A1"] = f"Partner Comparison — {payload['period']['label']}"
    s["A1"].font = c["TITLE_FONT"]; s.merge_cells("A1:H1")
    s["A2"] = f"Business: {biz}    Total shipments: {payload['total_shipments']}"
    s["A2"].font = c["LABEL_FONT"]; s.merge_cells("A2:H2")

    for col, h in enumerate(
        ["Courier", "Total", "Delivered", "Returned", "Pending",
         "Delivery %", "Return %", "Avg Cost ₹", "Revenue ₹", "Margin ₹", "COD ₹"], 1):
        cell = s.cell(row=4, column=col, value=h)
        cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
    for i, b in enumerate(payload["couriers"], 1):
        r = 4 + i
        s.cell(row=r, column=1,  value=b["courier_name"])
        s.cell(row=r, column=2,  value=b["total"])
        s.cell(row=r, column=3,  value=b["delivered"])
        s.cell(row=r, column=4,  value=b["returned"])
        s.cell(row=r, column=5,  value=b["pending"])
        s.cell(row=r, column=6,  value=f"{b['delivery_rate']}%")
        s.cell(row=r, column=7,  value=f"{b['return_rate']}%")
        s.cell(row=r, column=8,  value=round(b["avg_cost"]))
        s.cell(row=r, column=9,  value=round(b["revenue"]))
        s.cell(row=r, column=10, value=round(b["margin"]))
        s.cell(row=r, column=11, value=round(b["cod"]))
    for col in range(1, 12):
        s.column_dimensions[c["get_column_letter"](col)].width = 14

    # Rankings summary
    base = 5 + len(payload["couriers"]) + 1
    s.cell(row=base, column=1, value="Rankings").font = c["TITLE_FONT"]
    rk = payload["rankings"]
    rows_meta = [
        ("Best delivery rate",   rk["best_delivery"]),
        ("Worst delivery rate",  rk["worst_delivery"]),
        ("Lowest return rate",   rk["best_returns"]),
        ("Highest return rate",  rk["worst_returns"]),
        ("Cheapest courier",     rk["cheapest"]),
        ("Highest revenue",      rk["highest_revenue"]),
    ]
    for i, (lbl, val) in enumerate(rows_meta, 1):
        s.cell(row=base+i, column=1, value=lbl).font = c["LABEL_FONT"]
        s.cell(row=base+i, column=2, value=val or "—")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return _xlsx_response(buf, prefix="PartnerComparison", label=payload["period"]["label"])


def _build_recon_excel(payload, current_user) -> StreamingResponse:
    c = _xlsx_chrome()
    wb = c["Workbook"]()
    biz = current_user.get("shop_name") or current_user.get("name") or "—"

    s = wb.active
    s.title = "Reconciliation"
    s["A1"] = f"COD Reconciliation — {payload['period']['label']}"
    s["A1"].font = c["TITLE_FONT"]; s.merge_cells("A1:G1")
    s["A2"] = f"Business: {biz}"
    s["A2"].font = c["LABEL_FONT"]; s.merge_cells("A2:G2")

    t = payload["totals"]
    s["A3"] = (f"Delivered COD: {t['delivered_count']} (₹{t['delivered_amt']:.0f})    "
               f"Received: {t['received_count']} (₹{t['received_amt']:.0f})    "
               f"Pending: {t['pending_count']} (₹{t['pending_amt']:.0f})")
    s["A3"].font = c["LABEL_FONT"]; s.merge_cells("A3:G3")

    for col, h in enumerate(
        ["Courier", "Delivered #", "Delivered ₹", "Received #", "Received ₹",
         "Pending #", "Pending ₹"], 1):
        cell = s.cell(row=5, column=col, value=h)
        cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
    for i, b in enumerate(payload["couriers"], 1):
        r = 5 + i
        s.cell(row=r, column=1, value=b["courier_name"])
        s.cell(row=r, column=2, value=b["delivered_count"])
        s.cell(row=r, column=3, value=round(b["delivered_amt"]))
        s.cell(row=r, column=4, value=b["received_count"])
        s.cell(row=r, column=5, value=round(b["received_amt"]))
        s.cell(row=r, column=6, value=b["pending_count"])
        s.cell(row=r, column=7, value=round(b["pending_amt"]))
    for col in range(1, 8):
        s.column_dimensions[c["get_column_letter"](col)].width = 16

    # Pending COD list (delivered-but-not-received)
    if payload["pending"]:
        ws = wb.create_sheet(title="Pending COD")
        for col, h in enumerate(
            ["Delivered Date", "Tracking ID", "Order ID",
             "Customer", "Courier", "Amount ₹"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill, cell.font, cell.alignment = c["HEADER_FILL"], c["HEADER_FONT"], c["CENTER"]
        for i, r in enumerate(payload["pending"], 1):
            row = 1 + i
            ws.cell(row=row, column=1, value=r.get("delivered_at"))
            ws.cell(row=row, column=2, value=r.get("tracking_id"))
            ws.cell(row=row, column=3, value=r.get("order_id"))
            ws.cell(row=row, column=4, value=r.get("customer"))
            ws.cell(row=row, column=5, value=r.get("courier"))
            ws.cell(row=row, column=6, value=round(r.get("amount") or 0))
        for col in range(1, 7):
            ws.column_dimensions[c["get_column_letter"](col)].width = 18

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return _xlsx_response(buf, prefix="Reconciliation", label=payload["period"]["label"])
