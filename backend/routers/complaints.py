"""
India Post Complaint Management Router — Phase F7.0 (June 2026).

Adds a lightweight per-shipment complaint sub-record and a specialised
bulk Excel export that matches the strict India Post CBS complaint
upload template. Kept out of `shipments_write.py` / `shipments_export.py`
so the two large router files don't grow further.

Endpoints
─────────
    PATCH  /api/shipments/{id}/complaint            → save / update complaint
    DELETE /api/shipments/{id}/complaint            → clear complaint
    POST   /api/shipments/export-complaints         → bulk India Post Excel
                                                       (ZIP with 500-row chunks)
    POST   /api/shipments/complaint/ai-shorten      → LLM-rewrite Description
                                                       to ≤250-char English
                                                       (charges 0.5 AI credits
                                                       ONLY on success)

DB additions (all on the existing `shipments` collection, additive)
────────────────────────────────────────────────────────────────────
    complaint_created           : bool           (mirrors "has complaint" for fast filter)
    complaint_booking_date      : str  DD-MM-YYYY (India Post format)
    complaint_service_name      : str            SP_INLAND_PARCEL / SP_SPEED_POST_EMO / SP_BUSINESS_PARCEL / Other
    complaint_service_name_other: str            free-text when service = Other
    complaint_type              : str            Non-Delivery / Delayed Delivery / Damaged/Loss / Wrong Delivery
    complaint_description       : str
    complaint_status            : str            Open / In Progress / Resolved / Closed (UI only, not exported)
    complaint_created_at        : ISO datetime
    complaint_updated_at        : ISO datetime

Excel schema (India Post CBS bulk template — DO NOT reorder columns)
────────────────────────────────────────────────────────────────────
    A: Serial No
    B: Order No
    C: Article No (e.g. EA123456789IN)
    D: Booking Date
    E: Service Name
    F: Complaint Type
    G: Description

Filename pattern: IndiaPost_Complaint_DD-MM-YYYY_Part{N}.xlsx
Chunking: max 500 rows per file (India Post upload cap).
Wrapper: single ZIP so the client only saves one file.
"""
from __future__ import annotations

import io
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from pydantic import BaseModel, Field

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response


complaints_router = APIRouter(prefix="/api", tags=["complaints"])


# ─── Payloads ────────────────────────────────────────────────────────
class _ComplaintPayload(BaseModel):
    """Complaint fields the client can PATCH on a shipment."""
    booking_date:         Optional[str] = None   # DD-MM-YYYY (India Post format)
    service_name:         Optional[str] = None   # dropdown enum
    service_name_other:   Optional[str] = None   # free-text when service = "Other"
    complaint_type:       Optional[str] = None   # dropdown enum
    complaint_description: Optional[str] = None
    complaint_status:     Optional[str] = None   # dropdown enum


class _ExportComplaintsBody(BaseModel):
    """Filtered-list ID payload — identical shape to /shipments/export/*"""
    ids: Optional[List[str]] = None


class _AiShortenBody(BaseModel):
    """Payload for POST /shipments/complaint/ai-shorten.

    NOTE: MUST live at module scope. FastAPI/Pydantic v2 fail to
    recognise Pydantic models defined inside a function as request-body
    schemas and silently degrade them to query parameters. Keep this
    class outside init().
    """
    description: str


_VALID_SERVICES = {
    # Phase F7.3 (Jun-2026) — India Post service catalogue (verbatim
    # copy of the CBS complaint form dropdown). Order in the frontend
    # is defined separately; this set is only for validation.
    "Speed Post Parcel (Registered/Insured/COD)",
    "Speed Post Letters (Insured/COD)",
    "India Post Parcel- Contractual (Registered/Insured/COD)",
    "India Post Parcel-retail (Registered/Insured/COD)",
    "Letter (Registered/Insured/COD)",
    "Magazine Post",
    "Post Card/Book Post/Periodical Post/Registered Newspapers/Inland Letter Card/Ordinary letter",
    "Tariff /GST Related",
}
_DEFAULT_SERVICE = "Speed Post Parcel (Registered/Insured/COD)"
# India Post CBS official complaint types — DO NOT reword, DO NOT translate.
# The values below MUST match the official India Post complaint dropdown
# character-for-character (including casing / spacing) or the CBS bulk
# uploader rejects the row.
_VALID_TYPES = {
    "Delay in delivery",
    "Non delivery of article",
    "Abstraction of Contents",
    "Loss of article",
    "Non payment of COD Amount",
    "Damage of Article",
    "Fake/Non-updation of delivery remarks/Scans",
}
_VALID_STATUSES = {"Open", "In Progress", "Resolved", "Closed"}


def _normalise_booking_date(raw: Optional[str]) -> str:
    """Force DD-MM-YYYY string (India Post's only accepted format).

    Accepts multiple client conventions (ISO, DD/MM/YYYY, DD-MM-YYYY,
    YYYY-MM-DD) so the UI date picker doesn't have to be locale-aware.
    """
    if not raw:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    # Try common shapes in priority order.
    fmts = (
        "%d-%m-%Y", "%d/%m/%Y",
        "%Y-%m-%d", "%Y/%m/%d",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ",
    )
    for f in fmts:
        try:
            return datetime.strptime(s.split(".")[0].replace("Z", ""), f).strftime("%d-%m-%Y")
        except Exception:
            continue
    # Last-ditch: pass through untouched (client will see it and can retry).
    return s


def init() -> None:
    """Register complaint routes after server.py has initialised db + auth."""
    import logging
    _logger = logging.getLogger("routers.complaints")
    from server import (  # noqa: WPS433 — intentional late import
        db,
        get_current_user as _get_current_user,
    )

    # ═══════════════════════════════════════════════════════════════
    #                          PATCH / DELETE
    # ═══════════════════════════════════════════════════════════════
    @complaints_router.patch("/shipments/{shipment_id}/complaint")
    async def save_complaint(
        shipment_id: str,
        payload: _ComplaintPayload,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        # Validate the shipment exists and belongs to this user.
        ship = await db.shipments.find_one(
            {"id": shipment_id, "user_id": current_user["id"]},
            {"_id": 0},
        )
        if not ship:
            raise HTTPException(status_code=404, detail="Shipment not found")

        # Soft-validate enum inputs — reject only on hard mismatch so
        # the UI stays forgiving of accidental whitespace / casing.
        svc = (payload.service_name or "").strip()
        if svc and svc not in _VALID_SERVICES:
            raise HTTPException(
                status_code=422,
                detail=f"service_name must be one of {sorted(_VALID_SERVICES)}",
            )
        ctype = (payload.complaint_type or "").strip()
        if ctype and ctype not in _VALID_TYPES:
            raise HTTPException(
                status_code=422,
                detail=f"complaint_type must be one of {sorted(_VALID_TYPES)}",
            )
        cstat = (payload.complaint_status or "").strip() or "Open"
        if cstat not in _VALID_STATUSES:
            raise HTTPException(
                status_code=422,
                detail=f"complaint_status must be one of {sorted(_VALID_STATUSES)}",
            )

        now_iso = datetime.now(timezone.utc).isoformat()
        update: Dict[str, Any] = {
            "complaint_created": True,
            "complaint_booking_date": _normalise_booking_date(payload.booking_date),
            "complaint_service_name": svc or _DEFAULT_SERVICE,
            "complaint_service_name_other": (payload.service_name_other or "").strip(),
            "complaint_type": ctype,
            "complaint_description": (payload.complaint_description or "").strip(),
            "complaint_status": cstat,
            "complaint_updated_at": now_iso,
        }
        # First-time creation timestamp is set once and preserved on re-edits.
        if not ship.get("complaint_created_at"):
            update["complaint_created_at"] = now_iso

        await db.shipments.update_one(
            {"id": shipment_id, "user_id": current_user["id"]},
            {"$set": update},
        )
        # Return the freshly-persisted subset for optimistic UI.
        return {"ok": True, **update}

    @complaints_router.delete("/shipments/{shipment_id}/complaint")
    async def delete_complaint(
        shipment_id: str,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        ship = await db.shipments.find_one(
            {"id": shipment_id, "user_id": current_user["id"]},
            {"_id": 0, "id": 1},
        )
        if not ship:
            raise HTTPException(status_code=404, detail="Shipment not found")
        await db.shipments.update_one(
            {"id": shipment_id, "user_id": current_user["id"]},
            {"$set": {"complaint_created": False},
             "$unset": {
                 "complaint_booking_date": "",
                 "complaint_service_name": "",
                 "complaint_service_name_other": "",
                 "complaint_type": "",
                 "complaint_description": "",
                 "complaint_status": "",
                 "complaint_created_at": "",
                 "complaint_updated_at": "",
             }},
        )
        return {"ok": True}

    # ═══════════════════════════════════════════════════════════════
    #                    EXCEL BUILDER (chunked)
    # ═══════════════════════════════════════════════════════════════
    _CHUNK_SIZE = 500  # India Post CBS bulk-upload cap.
    # Column headers — MUST match the official India Post CBS complaint
    # template character-for-character. Do NOT reword or shorten.
    _COLUMNS = [
        "Serial No.",
        "Order / Transaction Number",
        "Article Number",
        "Booking Date",
        "Service Name",
        "Complaint Type",
        "Description",
    ]

    def _build_single_xlsx(rows: List[List[Any]], start_serial: int) -> bytes:
        """Build one India Post Complaint xlsx file for up to 500 rows."""
        # Lazy import — openpyxl is chunky and only needed for the export path.
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "IndiaPost Complaints"

        # Header row — plain (no background fill) to exactly mirror the
        # official India Post template. We keep only a subtle bold so
        # the header is still visually distinguishable in wide sheets.
        ws.append(_COLUMNS)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Data rows — Serial No is 1-based INSIDE THIS FILE (not across
        # files). India Post's uploader treats each part independently.
        for i, r in enumerate(rows, start=1):
            ws.append([i, *r])

        # Freeze header + widen readable columns.
        ws.freeze_panes = "A2"
        widths = [10, 28, 22, 14, 22, 42, 60]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def _fetch_complaint_rows(
        user_id: str, ids: Optional[List[str]] = None,
    ) -> List[List[Any]]:
        """Return raw row lists (Order No, Article No, Booking Date, Service Name,
        Complaint Type, Description) for every complaint-flagged shipment.

        Only shipments where `complaint_created == True` are exported —
        the "Complaint Created" filter on the Shipments screen sends the
        exact ID list, but we still double-check server-side to keep
        the endpoint safe from stale/broken filter payloads.
        """
        mongo_q: Dict[str, Any] = {
            "user_id": user_id,
            "complaint_created": True,
        }
        if ids:
            mongo_q["id"] = {"$in": list(ids)[:10_000]}
        projection = {
            "_id": 0,
            "id": 1, "order_id": 1, "tracking_id": 1,
            "complaint_booking_date": 1,
            "complaint_service_name": 1,
            "complaint_service_name_other": 1,
            "complaint_type": 1,
            "complaint_description": 1,
            "created_at": 1,
        }
        docs = await db.shipments.find(mongo_q, projection).sort(
            "complaint_created_at", -1,
        ).to_list(50_000)

        rows: List[List[Any]] = []
        for d in docs:
            svc = (d.get("complaint_service_name") or _DEFAULT_SERVICE).strip()
            # Phase F7.3 — "Other" free-text fallback removed from the
            # UI; if a legacy row still carries a `service_name_other`
            # we honour it so pre-existing complaints still export
            # meaningfully instead of silently dropping data.
            if svc == "Other":
                svc = (d.get("complaint_service_name_other") or "Other").strip()
            rows.append([
                d.get("order_id", "") or "",
                d.get("tracking_id", "") or "",
                d.get("complaint_booking_date", "") or "",
                svc,
                d.get("complaint_type", "") or "",
                d.get("complaint_description", "") or "",
            ])
        return rows

    # ═══════════════════════════════════════════════════════════════
    #                          EXPORT ROUTE
    # ═══════════════════════════════════════════════════════════════
    @complaints_router.post("/shipments/export-complaints")
    async def export_complaints(
        payload: _ExportComplaintsBody,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        rows = await _fetch_complaint_rows(current_user["id"], payload.ids)
        if not rows:
            raise HTTPException(
                status_code=404,
                detail="No complaint-flagged shipments to export.",
            )

        date_stamp = datetime.now().strftime("%d-%m-%Y")
        # Chunk into parts of 500 → build one xlsx per chunk → zip.
        parts: List[tuple[str, bytes]] = []
        for idx in range(0, len(rows), _CHUNK_SIZE):
            chunk = rows[idx: idx + _CHUNK_SIZE]
            part_no = (idx // _CHUNK_SIZE) + 1
            xlsx = _build_single_xlsx(chunk, start_serial=idx + 1)
            fname = f"IndiaPost_Complaint_{date_stamp}_Part{part_no}.xlsx"
            parts.append((fname, xlsx))

        # Single file → return the xlsx directly (nicer UX than a 1-file zip).
        if len(parts) == 1:
            fname, data = parts[0]
            return Response(
                content=data,
                media_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                headers={
                    "Content-Disposition": f'attachment; filename="{fname}"',
                    "X-Complaint-Parts": "1",
                    "X-Complaint-Total-Rows": str(len(rows)),
                },
            )

        # Multiple parts → wrap all xlsx into a single ZIP so the
        # client saves one file and gets N split-uploads inside.
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
            for fname, data in parts:
                z.writestr(fname, data)
        zip_buf.seek(0)
        zip_name = f"IndiaPost_Complaint_{date_stamp}.zip"
        return Response(
            content=zip_buf.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{zip_name}"',
                "X-Complaint-Parts": str(len(parts)),
                "X-Complaint-Total-Rows": str(len(rows)),
            },
        )

    # ═══════════════════════════════════════════════════════════════
    #                    AI SHORTEN — LLM rewrite
    # ═══════════════════════════════════════════════════════════════
    #
    # Contract:
    #   • Accepts ANY language (Gujarati / Hindi / Marathi / mixed).
    #   • Rewrites into professional English ≤ 250 characters.
    #   • Loops up to 3 times if the model overshoots 250 chars — each
    #     retry passes the previous output back with a stricter
    #     length reminder.
    #   • Charges 0.5 AI credits ONLY after a SUCCESSFUL rewrite
    #     (no charge on timeout / API error / empty response).
    #   • Every successful click is billed independently — there is no
    #     "one shorten per complaint" throttle. This matches the user
    #     spec: 1 click = 0.5, 2 clicks = 1.0, etc.
    _AI_SHORTEN_COST = 0.5

    _SHORTEN_SYSTEM = (
        "You are an India Post complaint rewriter. The user gives you "
        "a complaint description in ANY language (Gujarati, Hindi, "
        "Marathi, English, or mixed). Your job:\n"
        "  1. Auto-detect the language.\n"
        "  2. Translate everything into professional English.\n"
        "  3. Preserve EVERY important complaint fact (tracking IDs, "
        "     dates, article names, amounts, delivery attempts, "
        "     contact history, damage details, etc.).\n"
        "  4. Do NOT change the meaning. Do NOT drop information.\n"
        "  5. Remove ONLY redundant words, fillers and repetition.\n"
        "  6. Write in a short, clear, professional complaint tone "
        "     suitable for the India Post CBS Bulk Complaint upload.\n"
        "  7. HARD LIMIT: the final English output MUST be at most "
        "     250 characters (including spaces & punctuation).\n"
        "  8. Return ONLY the rewritten English complaint text. NO "
        "     preamble. NO markdown. NO quotes. NO labels."
    )

    @complaints_router.post("/shipments/complaint/ai-shorten")
    async def ai_shorten_complaint(
        payload: _AiShortenBody,
        current_user: Dict[str, Any] = Depends(_get_current_user),
    ):
        import os as _os
        import asyncio as _asyncio
        from wallet import (  # local import so wallet stays optional
            get_balance as _get_balance,
            _apply_delta as _apply_wallet_delta,
            record_history as _record_history,
        )

        raw = (payload.description or "").strip()
        if not raw:
            raise HTTPException(
                status_code=422,
                detail="Description is empty — nothing to shorten.",
            )
        # Guard against absurd inputs — India Post CBS won't accept
        # anything above a few KB anyway, and the LLM is metered by
        # token so we cap at 4 KB.
        if len(raw) > 4000:
            raise HTTPException(
                status_code=413,
                detail="Description too long to shorten (max 4000 chars).",
            )

        # ── Pre-flight: refuse if the wallet doesn't have 0.5 credits.
        # We intentionally check BEFORE calling the LLM so we never
        # burn tokens for a user who can't pay.
        bal_before = await _get_balance(db, current_user["id"])
        if bal_before < _AI_SHORTEN_COST - 1e-6:
            raise HTTPException(
                status_code=402,
                detail=(
                    f"Insufficient AI credits. AI Shorten needs "
                    f"{_AI_SHORTEN_COST} credits; you have "
                    f"{round(bal_before, 2)}. Top up from Wallet."
                ),
            )

        api_key = _os.getenv("EMERGENT_LLM_KEY", "").strip()
        if not api_key:
            raise HTTPException(
                status_code=503,
                detail="AI service is not configured (EMERGENT_LLM_KEY missing).",
            )

        # ─── LLM call with up to 3 length-shrinking retries ────────
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
        except Exception as e:  # pragma: no cover — package missing
            raise HTTPException(
                status_code=503,
                detail=f"AI runtime unavailable: {e}",
            )

        last_output: str = ""
        rewritten: Optional[str] = None
        _timeout = float(_os.getenv("AI_SHORTEN_TIMEOUT", "20.0"))
        # Provider/model tuned for short, deterministic transforms —
        # nano is cheap + fast, works fine for translate-and-compress.
        provider = _os.getenv("AI_SHORTEN_PROVIDER", "openai")
        model    = _os.getenv("AI_SHORTEN_MODEL",    "gpt-4.1-mini")

        try:
            for attempt in range(3):
                session_id = (
                    f"complaint-shorten-{current_user['id']}-{attempt}-"
                    f"{abs(hash(raw[:200]))}"
                )
                chat = (
                    LlmChat(
                        api_key=api_key,
                        session_id=session_id,
                        system_message=_SHORTEN_SYSTEM,
                    )
                    .with_model(provider, model)
                )
                if attempt == 0:
                    user_text = (
                        "Rewrite this complaint into professional "
                        "English ≤250 characters:\n\n" + raw
                    )
                else:
                    # Retry — feed both the original and the previous
                    # (too-long) attempt so the model can compress
                    # further without losing facts.
                    user_text = (
                        f"Your previous output was {len(last_output)} "
                        "characters — TOO LONG. Rewrite the ORIGINAL "
                        "complaint below in ≤250 characters of "
                        "professional English WITHOUT losing any "
                        f"important fact:\n\nORIGINAL:\n{raw}\n\n"
                        f"PREVIOUS ATTEMPT:\n{last_output}\n\n"
                        "Now output the ≤250-char English version:"
                    )
                out = await _asyncio.wait_for(
                    chat.send_message(UserMessage(text=user_text)),
                    timeout=_timeout,
                )
                out = (out or "").strip().strip('"').strip("'")
                # Strip common leading labels ("English: ", "Rewritten: ").
                for prefix in ("English:", "Rewritten:", "Output:", "Complaint:"):
                    if out.lower().startswith(prefix.lower()):
                        out = out[len(prefix):].strip()
                last_output = out
                if out and len(out) <= 250:
                    rewritten = out
                    break
        except _asyncio.TimeoutError:
            raise HTTPException(
                status_code=504,
                detail="AI Shorten timed out. Please try again.",
            )
        except HTTPException:
            raise
        except Exception as e:
            _logger.warning("AI Shorten LLM error: %s", e)
            raise HTTPException(
                status_code=502,
                detail=(
                    "AI Shorten failed. No credits were deducted. "
                    "Please try again."
                ),
            )

        if not rewritten:
            # Model couldn't fit 250 chars even after 3 tries — do NOT
            # deduct credits (per spec: no charge unless a valid
            # rewritten description is generated).
            raise HTTPException(
                status_code=502,
                detail=(
                    "AI Shorten couldn't fit the complaint under 250 "
                    "characters. No credits were deducted."
                ),
            )

        # ── Success → deduct 0.5 credits and log history ────────────
        uid = current_user["id"]
        w = await _apply_wallet_delta(db, uid, used_delta=_AI_SHORTEN_COST)
        bal_after = float(w.get("remaining_credits", 0.0))
        await _record_history(
            db, uid,
            credits=-_AI_SHORTEN_COST,
            ctype="ai_processing",
            description="AI Shorten · India Post Complaint",
            balance_after=bal_after,
        )
        return {
            "ok": True,
            "rewritten": rewritten,
            "credits_deducted": _AI_SHORTEN_COST,
            "balance_after": bal_after,
            "chars": len(rewritten),
        }

    _logger.info(
        "complaints router mounted: 4 endpoints "
        "(PATCH + DELETE /shipments/{id}/complaint, "
        "POST /shipments/export-complaints, "
        "POST /shipments/complaint/ai-shorten)"
    )
