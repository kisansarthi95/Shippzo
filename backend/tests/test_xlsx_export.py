"""
Iteration 34 — XLSX export endpoint tests (Gujarati/Hindi encoding fix).

Tests T1-T6 as defined in the review request:
- T1: XLSX endpoint contract (status, content-type, disposition, PK magic)
- T2: XLSX file integrity (sheet name, column count, header row, freeze pane)
- T3: Unicode preservation (Gujarati customer name round-trip - THE CRITICAL BUG)
- T4: Filter-aware export via POST with ids
- T5: Empty-ID list path (nonexistent id → header-only workbook)
- T6: Regression — CSV export still emits UTF-8 BOM

Note: T7 (re-run smart_search + sheet_orders_case_fix) is executed separately
via pytest command in the test harness — not embedded here.
"""
import os
from io import BytesIO
from typing import Any, Dict

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # Fall back — read directly from frontend/.env
    with open("/app/frontend/.env") as _fh:
        for _line in _fh:
            if _line.startswith("EXPO_PUBLIC_BACKEND_URL="):
                BASE_URL = _line.split("=", 1)[1].strip().rstrip("/")
                break

ADMIN_EMAIL = "admin@test.com"
ADMIN_PASSWORD = "Admin@12345"

XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
ZIP_MAGIC = b"PK\x03\x04"


# ────────────────────────── Fixtures ──────────────────────────

@pytest.fixture(scope="module")
def admin_token() -> str:
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=30,
    )
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    tok = r.json().get("token")
    assert tok, f"no token in login response: {r.json()}"
    return tok


@pytest.fixture(scope="module")
def auth_headers(admin_token: str) -> Dict[str, str]:
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="module")
def gujarati_shipment(auth_headers) -> Dict[str, Any]:
    """Create a fresh shipment with Gujarati customer name + city.
    Yields the created shipment dict; deletes it after the module runs.
    """
    import uuid
    payload = {
        "tracking_id": f"TEST_XLSX_{uuid.uuid4().hex[:8].upper()}",
        "customer_name": "સોનુ કુશવાહ",
        "customer_phone": "9876500001",
        "address_line1": "TEST_XLSX Address Line 1",
        "city": "ભાવનગર",
        "state": "Gujarat",
        "pincode": "364001",
        "items": ["TEST_XLSX_ITEM"],
        "amount": 499.0,
        "payment_mode": "COD",
        "payment_type": "COD",
    }
    r = requests.post(
        f"{BASE_URL}/api/shipments",
        json=payload,
        headers=auth_headers,
        timeout=30,
    )
    assert r.status_code in (200, 201), (
        f"create failed: {r.status_code} {r.text}"
    )
    created = r.json()
    assert created.get("id"), f"no id on created shipment: {created}"

    yield created

    # Teardown — clean up the created shipment.
    try:
        requests.delete(
            f"{BASE_URL}/api/shipments/{created['id']}",
            headers=auth_headers,
            timeout=15,
        )
    except Exception:
        pass


# ────────────────────────── Test Cases ──────────────────────────


# ── T1: XLSX endpoint contract ─────────────────────────────────
class TestT1XlsxContract:

    def test_get_xlsx_returns_200_and_correct_headers(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/shipments/export/xlsx",
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200, f"got {r.status_code}: {r.text[:200]}"
        assert r.headers.get("content-type", "").lower().startswith(
            XLSX_MIME
        ), f"unexpected content-type: {r.headers.get('content-type')}"
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower(), f"missing attachment: {cd}"
        assert ".xlsx" in cd.lower(), f"missing .xlsx in disposition: {cd}"

    def test_xlsx_body_starts_with_zip_magic(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/shipments/export/xlsx",
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        assert r.content[:4] == ZIP_MAGIC, (
            f"expected PK\\x03\\x04, got {r.content[:4]!r}"
        )


# ── T2: XLSX file integrity ────────────────────────────────────
class TestT2XlsxIntegrity:

    def test_workbook_structure(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/shipments/export/xlsx",
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert wb.sheetnames == ["Shipments"], (
            f"unexpected sheets: {wb.sheetnames}"
        )
        ws = wb["Shipments"]
        # Header row
        header = [c.value for c in ws[1]]
        assert len(header) == 45, f"expected 45 columns, got {len(header)}"
        assert header[:5] == [
            "Shipment ID", "Tracking ID", "Master Order ID",
            "Order ID", "AWB Number",
        ], f"unexpected first 5 header cells: {header[:5]}"
        # Freeze pane
        assert ws.freeze_panes == "A2", (
            f"expected freeze_panes 'A2', got {ws.freeze_panes!r}"
        )


# ── T3: Unicode preservation (THE CRITICAL BUG) ────────────────
class TestT3UnicodeRoundTrip:

    def test_gujarati_customer_name_preserved(
        self, auth_headers, gujarati_shipment
    ):
        sid = gujarati_shipment["id"]
        r = requests.post(
            f"{BASE_URL}/api/shipments/export/xlsx",
            json={"ids": [sid]},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200, f"got {r.status_code}: {r.text[:200]}"
        assert r.content[:4] == ZIP_MAGIC

        wb = load_workbook(BytesIO(r.content))
        ws = wb["Shipments"]
        header = [c.value for c in ws[1]]
        col_shipment_id = header.index("Shipment ID")
        col_customer_name = header.index("Customer Name")
        col_city = header.index("City")

        # Find the row matching this shipment id
        found_row = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_shipment_id] == sid:
                found_row = row
                break

        assert found_row is not None, (
            f"row with id={sid} not found in XLSX export"
        )
        customer_cell = found_row[col_customer_name]
        city_cell = found_row[col_city]
        assert customer_cell == "સોનુ કુશવાહ", (
            f"customer_name mismatch — got {customer_cell!r} "
            f"(bytes: {customer_cell.encode('utf-8') if customer_cell else b''!r})"
        )
        assert city_cell == "ભાવનગર", (
            f"city mismatch — got {city_cell!r}"
        )


# ── T4: Filter-aware export ────────────────────────────────────
class TestT4FilterAware:

    def test_post_with_two_ids_returns_two_data_rows(self, auth_headers):
        # First list some existing shipments to pick 2 valid ids.
        r_list = requests.get(
            f"{BASE_URL}/api/shipments?limit=10",
            headers=auth_headers,
            timeout=30,
        )
        assert r_list.status_code == 200
        shipments = r_list.json()
        if len(shipments) < 2:
            pytest.skip(f"need >=2 shipments in db, found {len(shipments)}")
        id_a = shipments[0]["id"]
        id_b = shipments[1]["id"]

        r = requests.post(
            f"{BASE_URL}/api/shipments/export/xlsx",
            json={"ids": [id_a, id_b]},
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        ws = wb["Shipments"]
        # header + 2 data rows == max_row 3
        assert ws.max_row == 3, (
            f"expected 3 rows (1 header + 2 data), got {ws.max_row}"
        )
        header = [c.value for c in ws[1]]
        col_sid = header.index("Shipment ID")
        ids_in_col = {
            row[col_sid]
            for row in ws.iter_rows(min_row=2, values_only=True)
        }
        assert id_a in ids_in_col and id_b in ids_in_col, (
            f"expected {id_a}, {id_b} in first col, got {ids_in_col}"
        )


# ── T5: Empty-ID list path ─────────────────────────────────────
class TestT5NonexistentIdOnly:

    def test_nonexistent_id_returns_200_header_only(self, auth_headers):
        r = requests.post(
            f"{BASE_URL}/api/shipments/export/xlsx",
            json={"ids": ["nonexistent-id-xxx"]},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, f"got {r.status_code}: {r.text[:200]}"
        wb = load_workbook(BytesIO(r.content))
        ws = wb["Shipments"]
        assert ws.max_row == 1, (
            f"expected only header row, got max_row={ws.max_row}"
        )


# ── T6: Regression — CSV export still works ───────────────────
class TestT6CsvRegression:

    def test_csv_export_has_utf8_bom_and_charset(self, auth_headers):
        r = requests.get(
            f"{BASE_URL}/api/shipments/export/csv",
            headers=auth_headers,
            timeout=60,
        )
        assert r.status_code == 200
        ct = r.headers.get("content-type", "").lower()
        assert "text/csv" in ct, f"unexpected content-type: {ct}"
        # UTF-8 BOM prefix
        assert r.content[:3] == b"\xef\xbb\xbf", (
            f"expected UTF-8 BOM prefix, got {r.content[:3]!r}"
        )
        # Rows > 0 — count newlines after header
        text = r.content.decode("utf-8-sig")
        lines = [l for l in text.splitlines() if l.strip()]
        assert len(lines) >= 2, (
            f"expected header + at least 1 row, got {len(lines)} lines"
        )
