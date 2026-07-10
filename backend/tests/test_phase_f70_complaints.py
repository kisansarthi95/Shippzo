"""
Phase F7.0 — India Post Complaint Management tests.

Coverage:
  • Complaint CRUD  (PATCH / DELETE /api/shipments/{id}/complaint)
  • Validation      (invalid enums, 404 for wrong owner / missing shipment)
  • Date normalise  (multiple input shapes → DD-MM-YYYY)
  • Excel export   (POST /api/shipments/export-complaints)
      – single-file xlsx
      – multi-part ZIP (>500 rows)
      – exact column header order + no "Complaint Status" column
      – Serial No 1-based per part
      – Service Name = free-text when service == "Other"
      – payload.ids filter
"""
import io
import os
import zipfile
import pytest
import requests

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "").rstrip("/")
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL not set"

ADMIN_EMAIL = "admin@test.com"
ADMIN_PASS = "Admin@12345"
USER2_EMAIL = "user2@test.com"
USER2_PASS = "User@12345"

EXCEL_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
EXPECTED_COLUMNS = [
    "Serial No.",
    "Order / Transaction Number",
    "Article Number",
    "Booking Date",
    "Service Name",
    "Complaint Type",
    "Description",
]


# ─── Fixtures ────────────────────────────────────────────────────────
@pytest.fixture(scope="session")
def admin_headers():
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASS},
        timeout=15,
    )
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    tok = r.json()["token"]
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="session")
def user2_headers():
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": USER2_EMAIL, "password": USER2_PASS},
        timeout=15,
    )
    assert r.status_code == 200, f"User2 login failed: {r.status_code} {r.text}"
    tok = r.json()["token"]
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="session")
def admin_shipments(admin_headers):
    r = requests.get(f"{BASE_URL}/api/shipments", headers=admin_headers, timeout=20)
    assert r.status_code == 200
    ships = r.json()
    assert isinstance(ships, list) and len(ships) >= 2, "Need >=2 admin shipments"
    return ships


@pytest.fixture(scope="session")
def user2_shipment(user2_headers):
    r = requests.get(f"{BASE_URL}/api/shipments", headers=user2_headers, timeout=20)
    assert r.status_code == 200
    ships = r.json()
    if not ships:
        pytest.skip("user2 has no shipments")
    return ships[0]


# ─── Cleanup helper ──────────────────────────────────────────────────
def _clear_complaint(shipment_id, headers):
    requests.delete(
        f"{BASE_URL}/api/shipments/{shipment_id}/complaint",
        headers=headers, timeout=15,
    )


# ═════════════════════ Complaint CRUD ════════════════════════════════
class TestComplaintCRUD:
    def test_patch_valid_complaint_persists(self, admin_headers, admin_shipments):
        sid = admin_shipments[0]["id"]
        _clear_complaint(sid, admin_headers)
        payload = {
            "booking_date": "05-07-2026",
            "service_name": "SP_INLAND_PARCEL",
            "complaint_type": "Delay in delivery",
            "complaint_description": "Package not delivered after 10 days",
            "complaint_status": "Open",
        }
        r = requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers, json=payload, timeout=15,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["ok"] is True
        assert body["complaint_created"] is True
        assert body["complaint_booking_date"] == "05-07-2026"
        assert body["complaint_service_name"] == "SP_INLAND_PARCEL"
        assert body["complaint_type"] == "Delay in delivery"
        assert body["complaint_status"] == "Open"

        # GET verifies persistence
        g = requests.get(
            f"{BASE_URL}/api/shipments/{sid}", headers=admin_headers, timeout=15,
        )
        assert g.status_code == 200
        doc = g.json()
        assert doc.get("complaint_created") is True
        assert doc.get("complaint_description") == "Package not delivered after 10 days"

    @pytest.mark.parametrize(
        "field,val",
        [
            ("service_name", "INVALID_SVC"),
            ("complaint_type", "FooBar"),
            ("complaint_status", "Unknown"),
        ],
    )
    def test_invalid_enum_rejected(self, admin_headers, admin_shipments, field, val):
        sid = admin_shipments[0]["id"]
        base = {
            "booking_date": "05-07-2026",
            "service_name": "SP_INLAND_PARCEL",
            "complaint_type": "Delay in delivery",
            "complaint_description": "x",
            "complaint_status": "Open",
        }
        base[field] = val
        r = requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers, json=base, timeout=15,
        )
        assert r.status_code == 422, f"expected 422 for {field}={val}, got {r.status_code}"

    def test_service_other_with_free_text(self, admin_headers, admin_shipments):
        sid = admin_shipments[0]["id"]
        r = requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers,
            json={
                "booking_date": "05-07-2026",
                "service_name": "Other",
                "service_name_other": "SP_LOGISTICS_POST",
                "complaint_type": "Delay in delivery",
                "complaint_description": "custom",
                "complaint_status": "Open",
            }, timeout=15,
        )
        assert r.status_code == 200
        body = r.json()
        assert body["complaint_service_name"] == "Other"
        assert body["complaint_service_name_other"] == "SP_LOGISTICS_POST"

    @pytest.mark.parametrize(
        "raw", ["2026-07-05", "05/07/2026", "05-07-2026"],
    )
    def test_date_normalisation(self, admin_headers, admin_shipments, raw):
        sid = admin_shipments[0]["id"]
        r = requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers,
            json={
                "booking_date": raw,
                "service_name": "SP_INLAND_PARCEL",
                "complaint_type": "Delay in delivery",
                "complaint_description": "d",
                "complaint_status": "Open",
            }, timeout=15,
        )
        assert r.status_code == 200
        assert r.json()["complaint_booking_date"] == "05-07-2026", raw

    def test_patch_missing_shipment_returns_404(self, admin_headers):
        r = requests.patch(
            f"{BASE_URL}/api/shipments/does-not-exist-uuid/complaint",
            headers=admin_headers,
            json={"service_name": "SP_INLAND_PARCEL", "complaint_type": "Delay in delivery",
                  "complaint_description": "x", "complaint_status": "Open"},
            timeout=15,
        )
        assert r.status_code == 404

    def test_patch_other_users_shipment_returns_404(
        self, admin_headers, user2_headers, user2_shipment,
    ):
        sid = user2_shipment["id"]
        r = requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers,   # wrong user
            json={"service_name": "SP_INLAND_PARCEL", "complaint_type": "Delay in delivery",
                  "complaint_description": "x", "complaint_status": "Open"},
            timeout=15,
        )
        assert r.status_code == 404

    def test_delete_clears_complaint(self, admin_headers, admin_shipments):
        sid = admin_shipments[0]["id"]
        # ensure there's something to clear
        requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers,
            json={"booking_date": "05-07-2026", "service_name": "SP_INLAND_PARCEL",
                  "complaint_type": "Delay in delivery", "complaint_description": "d",
                  "complaint_status": "Open"},
            timeout=15,
        )
        d = requests.delete(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers, timeout=15,
        )
        assert d.status_code == 200
        g = requests.get(
            f"{BASE_URL}/api/shipments/{sid}", headers=admin_headers, timeout=15,
        )
        doc = g.json()
        assert doc.get("complaint_created") is False
        for k in (
            "complaint_booking_date", "complaint_service_name",
            "complaint_type", "complaint_description", "complaint_status",
        ):
            v = doc.get(k)
            assert v in (None, "", 0), f"{k} not cleared: {v!r}"


# ══════════════════════ Excel Export ═════════════════════════════════
class TestComplaintExport:
    def test_export_no_complaints_returns_404(self, admin_headers, admin_shipments):
        # First clear all complaints
        for s in admin_shipments:
            _clear_complaint(s["id"], admin_headers)
        r = requests.post(
            f"{BASE_URL}/api/shipments/export-complaints",
            headers=admin_headers, json={}, timeout=20,
        )
        assert r.status_code == 404

    def test_export_single_complaint_returns_xlsx(
        self, admin_headers, admin_shipments,
    ):
        sid = admin_shipments[0]["id"]
        # Setup: one complaint
        requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers,
            json={
                "booking_date": "05-07-2026",
                "service_name": "SP_INLAND_PARCEL",
                "complaint_type": "Delay in delivery",
                "complaint_description": "single-file test",
                "complaint_status": "Open",
            }, timeout=15,
        )
        r = requests.post(
            f"{BASE_URL}/api/shipments/export-complaints",
            headers=admin_headers, json={}, timeout=30,
        )
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith(EXCEL_MIME)
        assert r.headers.get("X-Complaint-Parts") == "1"
        cd = r.headers.get("content-disposition", "")
        assert "IndiaPost_Complaint_" in cd and "_Part1.xlsx" in cd

        # Verify columns + no Complaint Status
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header == EXPECTED_COLUMNS, header
        assert "Complaint Status" not in header
        # Serial No starts at 1
        assert ws.cell(row=2, column=1).value == 1
        _clear_complaint(sid, admin_headers)

    def test_export_multi_part_zip(self, admin_headers, admin_shipments):
        """501+ complaints → ZIP with multiple Part files, ≤500 rows each.

        Uses direct Mongo insertion for speed — creating 500+ shipments
        through the public API would take several minutes each run.
        """
        import asyncio
        import uuid
        from datetime import datetime, timezone
        from motor.motor_asyncio import AsyncIOMotorClient

        # We need the actual admin user_id — look it up
        me = requests.get(
            f"{BASE_URL}/api/auth/me", headers=admin_headers, timeout=15,
        ).json()
        user_id = me["id"]

        mongo_url = os.environ.get("MONGO_URL") or "mongodb://localhost:27017"
        db_name = os.environ.get("DB_NAME") or "test_database"

        async def _seed_bulk(n):
            cli = AsyncIOMotorClient(mongo_url)
            db = cli[db_name]
            now = datetime.now(timezone.utc).isoformat()
            docs = []
            ids = []
            for i in range(n):
                sid = f"TESTF70-{uuid.uuid4()}"
                ids.append(sid)
                docs.append({
                    "id": sid, "user_id": user_id,
                    "tracking_id": f"TESTF70{i:04d}",
                    "customer_name": f"TEST_F70 {i}",
                    "status": "Pending",
                    "created_at": now,
                    "amount": 1.0,
                    "order_id": f"OF70-{i}",
                    "complaint_created": True,
                    "complaint_booking_date": "05-07-2026",
                    "complaint_service_name": "SP_INLAND_PARCEL",
                    "complaint_type": "Delay in delivery",
                    "complaint_description": "bulk-seed",
                    "complaint_status": "Open",
                    "complaint_created_at": now,
                    "complaint_updated_at": now,
                })
            await db.shipments.insert_many(docs)
            cli.close()
            return ids

        async def _cleanup(ids):
            cli = AsyncIOMotorClient(mongo_url)
            db = cli[db_name]
            await db.shipments.delete_many({"id": {"$in": ids}})
            cli.close()

        # Clear any existing complaint flags first to have deterministic count
        for s in admin_shipments:
            _clear_complaint(s["id"], admin_headers)

        created_ids = asyncio.run(_seed_bulk(510))
        try:
            r = requests.post(
                f"{BASE_URL}/api/shipments/export-complaints",
                headers=admin_headers, json={}, timeout=60,
            )
            assert r.status_code == 200
            parts = int(r.headers.get("X-Complaint-Parts", "1"))
            total = int(r.headers.get("X-Complaint-Total-Rows", "0"))
            assert parts >= 2, f"expected >=2 parts, got {parts}"
            assert total >= 501
            assert r.headers.get("content-type", "").startswith("application/zip")
            # Verify zip contents
            zf = zipfile.ZipFile(io.BytesIO(r.content))
            names = zf.namelist()
            assert any("Part1" in n for n in names)
            assert any("Part2" in n for n in names)
            # Verify each part ≤500 rows and Serial 1-based per part
            from openpyxl import load_workbook
            for n in names:
                with zf.open(n) as fh:
                    wb = load_workbook(io.BytesIO(fh.read()))
                    ws = wb.active
                    header = [c.value for c in ws[1]]
                    assert header == EXPECTED_COLUMNS
                    data_rows = ws.max_row - 1
                    assert data_rows <= 500
                    assert ws.cell(row=2, column=1).value == 1
        finally:
            # cleanup
            asyncio.run(_cleanup(created_ids))
            for s in admin_shipments:
                _clear_complaint(s["id"], admin_headers)

    def test_export_service_other_uses_free_text(
        self, admin_headers, admin_shipments,
    ):
        sid = admin_shipments[0]["id"]
        requests.patch(
            f"{BASE_URL}/api/shipments/{sid}/complaint",
            headers=admin_headers,
            json={"booking_date": "05-07-2026",
                  "service_name": "Other",
                  "service_name_other": "SP_LOGISTICS_POST",
                  "complaint_type": "Delay in delivery",
                  "complaint_description": "other-test",
                  "complaint_status": "Open"},
            timeout=15,
        )
        r = requests.post(
            f"{BASE_URL}/api/shipments/export-complaints",
            headers=admin_headers, json={"ids": [sid]}, timeout=30,
        )
        assert r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Service Name column is E (index 5)
        svc_cell = ws.cell(row=2, column=5).value
        assert svc_cell == "SP_LOGISTICS_POST", (
            f"expected free-text service name, got {svc_cell!r}"
        )
        _clear_complaint(sid, admin_headers)

    def test_export_ids_filter(self, admin_headers, admin_shipments):
        s1, s2 = admin_shipments[0]["id"], admin_shipments[1]["id"]
        for sid, tag in ((s1, "A"), (s2, "B")):
            requests.patch(
                f"{BASE_URL}/api/shipments/{sid}/complaint",
                headers=admin_headers,
                json={"booking_date": "05-07-2026",
                      "service_name": "SP_INLAND_PARCEL",
                      "complaint_type": "Delay in delivery",
                      "complaint_description": f"filter-{tag}",
                      "complaint_status": "Open"},
                timeout=15,
            )
        # Only ask for s1
        r = requests.post(
            f"{BASE_URL}/api/shipments/export-complaints",
            headers=admin_headers, json={"ids": [s1]}, timeout=30,
        )
        assert r.status_code == 200
        assert r.headers.get("X-Complaint-Total-Rows") == "1"
        # cleanup
        _clear_complaint(s1, admin_headers)
        _clear_complaint(s2, admin_headers)
